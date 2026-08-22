# -*- coding: utf-8 -*-
"""2024 CUMCM Problem A 第四问求解脚本。

问题: 盘入螺线的螺距为 1.7 m，盘出螺线与盘入螺线关于螺线中心呈中心对称，
舞龙队在半径 R=4.5 m 的调头空间内完成调头。调头路径是由两段圆弧相切连接而成的
S 形曲线，前一段圆弧的半径是后一段的 2 倍，且与盘入、盘出螺线均相切。
(1) 讨论能否调整圆弧(仍保持各部分相切)使调头曲线变短；
(2) 龙头前把手速度恒为 1 m/s，以调头开始时间为零时刻，给出 -100..100 s
    每秒舞龙队全部把手的位置与速度，并在论文中给出 -100、-50、0、50、100 s
    时龙头前把手、第1/51/101/151/201节龙身前把手与龙尾后把手的位置和速度。

方法: 与策略方案第 4 节一致。由盘入螺线与调头边界圆求入口 A 与单位切向 T，
再由中心对称得出口 B=-A；双圆弧圆心分别位于 A、B 两侧的法线上，两圆外切条件
唯一确定 Q=R1+R2；R1:R2=2:1 给出具体半径。以沿运动方向的弧长坐标 s 统一参数化
复合路径(盘入螺线、第一圆弧、第二圆弧、盘出螺线)，逐节求解弦长约束得到 224 个
把手的路径坐标，再按弦长约束隐式求导递推速度。

输出:
  result4.xlsx                  -100..100 s 全量位置/速度、论文表5/表6、路径几何
  表5_表6_关键结果.txt          关键几何参数、不变性论证、数值校验与论文表
  调头路径示意图.png            盘入螺线 + S形双圆弧 + 盘出螺线 + A/J/B 切点
  调头全过程形态图.png          -100/-50/0/50/100 s 全龙形态
"""
import json
import math
import os

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import Circle
from openpyxl import Workbook

from common_model import (
    N_BOARDS,
    N_HANDLES,
    handle_distance,
    selected_handle_indices,
    spiral_b,
    TurnPath,
    path_state,
    constraint_residual,
)

plt.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei"]
plt.rcParams["axes.unicode_minus"] = False

BASE = os.path.dirname(os.path.abspath(__file__))
PITCH = 1.7
TURN_RADIUS = 4.5
RATIO = 2.0
TIMES = list(range(-100, 101))
SHOW_TIMES = [-100, -50, 0, 50, 100]
IDX = selected_handle_indices()


def handle_name(i: int) -> str:
    if i == 0:
        return "龙头"
    if 1 <= i <= 221:
        return f"第{i}节龙身"
    if i == 222:
        return "龙尾"
    return "龙尾（后）"


def fmt6(value: float) -> float:
    """保留 6 位小数，并把 -0.000000 归一为 0.000000。"""
    rounded = round(float(value), 6)
    return 0.0 if rounded == 0.0 else rounded


def tangent_angle(t1: np.ndarray, t2: np.ndarray) -> float:
    cross = float(t1[0] * t2[1] - t1[1] * t2[0])
    dot = float(t1 @ t2)
    return math.degrees(math.atan2(cross, dot))


def sample_path(path: TurnPath, s0: float, s1: float, n: int = 400) -> np.ndarray:
    ss = np.linspace(s0, s1, n)
    return np.array([path.point_tangent(float(s))[0] for s in ss])


def compute_states(path: TurnPath) -> dict:
    """逐秒计算 -100..100 s 全部把手状态。"""
    states = {}
    for t in TIMES:
        _, points, tangents, speeds = path_state(path, float(t), 1.0)
        states[t] = {"points": points, "tangents": tangents, "speeds": speeds}
    return states


def geometry_checks(path: TurnPath) -> dict:
    report = path.geometry_report()
    oo_distance = float(np.linalg.norm(path.o2 - path.o1))
    circle_tangency_error = abs(oo_distance - (path.r1 + path.r2))

    tangent_in, tangent_out = path.point_tangent(-1e-9)[1], path.point_tangent(1e-9)[1]
    angle_a = tangent_angle(tangent_in, tangent_out)

    before_j = path.point_tangent(path.length1 - 1e-9)[1]
    after_j = path.point_tangent(path.length1 + 1e-9)[1]
    angle_j = tangent_angle(before_j, after_j)

    before_b = path.point_tangent(path.turn_length - 1e-9)[1]
    after_b = path.point_tangent(path.turn_length + 1e-9)[1]
    angle_b = tangent_angle(before_b, after_b)

    point_a_error = float(np.linalg.norm(path.a - np.array(report["A"])))
    point_b_error = float(np.linalg.norm((-path.a) - np.array(report["E"])))
    return {
        "theta_a": report["theta_a"],
        "A": report["A"],
        "B": report["E"],
        "T": path.tangent,
        "N": path.normal,
        "Q": path.radius_sum,
        "R1": path.r1,
        "R2": path.r2,
        "O1": path.o1,
        "O2": path.o2,
        "J": path.c,
        "alpha": path.angle1,
        "angle2": path.angle2,
        "length1": path.length1,
        "length2": path.length2,
        "turn_length": path.turn_length,
        "max_radius_on_turn": report["max_radius_on_turn"],
        "junction_position_error": report["junction_position_error"],
        "circle_tangency_error": circle_tangency_error,
        "tangent_angle_at_A_deg": angle_a,
        "tangent_angle_at_J_deg": angle_j,
        "tangent_angle_at_B_deg": angle_b,
        "point_a_error": point_a_error,
        "point_b_error": point_b_error,
    }


def invariance_sweep() -> list:
    """保持 A、B 与两端切向不变，仅改变半径比例 λ，验证总长不变。"""
    rows = []
    for ratio in (0.5, 0.75, 1.0, 1.5, 2.0, 3.0, 4.0):
        p = TurnPath(PITCH, TURN_RADIUS, ratio)
        rep = p.geometry_report()
        rows.append({
            "ratio": ratio,
            "R1": p.r1,
            "R2": p.r2,
            "turn_length": p.turn_length,
            "max_radius_on_turn": rep["max_radius_on_turn"],
        })
    return rows


def finite_difference_check(path: TurnPath, states: dict, dt: float = 1e-5) -> dict:
    """解析速度递推与中心有限差分速度对照。"""
    worst_rel = 0.0
    worst_abs = 0.0
    for t in range(-100, 101, 10):
        _, p_plus, _, _ = path_state(path, t + dt, 1.0)
        _, p_minus, _, _ = path_state(path, t - dt, 1.0)
        speeds = states[t]["speeds"]
        fd = np.linalg.norm(p_plus - p_minus, axis=1) / (2.0 * dt)
        rel = np.max(np.abs(fd - speeds) / speeds)
        abs_err = np.max(np.abs(fd - speeds))
        worst_rel = max(worst_rel, float(rel))
        worst_abs = max(worst_abs, float(abs_err))
    return {"relative": worst_rel, "absolute": worst_abs}


def reference_comparison(states: dict) -> float:
    ref_path = os.path.join(BASE, "..", "..", "24A题_独立复现", "independent_results.json")
    if not os.path.exists(ref_path):
        return None
    with open(ref_path, encoding="utf-8") as fp:
        ref = json.load(fp)["problem4"]
    worst = 0.0
    for key, val in ref.items():
        t = int(key)
        pos = states[t]["points"]
        spd = states[t]["speeds"]
        for j, i in enumerate(IDX):
            worst = max(worst, abs(pos[i][0] - val["positions"][j][0]),
                        abs(pos[i][1] - val["positions"][j][1]),
                        abs(spd[i] - val["speeds"][j]))
    return float(worst)


def chord_residual_max(states: dict) -> float:
    return max(constraint_residual(states[t]["points"]) for t in TIMES)


def write_result4(path: TurnPath, geom: dict, states: dict, sweep: list,
                  fd: dict, ref_err) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "位置"
    header = ["t (s)"]
    for i in range(N_HANDLES):
        header += [handle_name(i) + "x (m)", handle_name(i) + "y (m)"]
    ws.append(header)
    for t in TIMES:
        row = [t]
        for i in range(N_HANDLES):
            row += [fmt6(states[t]["points"][i][0]), fmt6(states[t]["points"][i][1])]
        ws.append(row)

    ws = wb.create_sheet("速度")
    ws.append(["t (s)"] + [handle_name(i) + " (m/s)" for i in range(N_HANDLES)])
    for t in TIMES:
        ws.append([t] + [fmt6(states[t]["speeds"][i]) for i in range(N_HANDLES)])

    def paper_sheet(title, rows, cols):
        ws = wb.create_sheet(title)
        ws.append(["部位"] + ["%d s" % c for c in cols])
        for r in rows:
            ws.append(r)

    table5 = []
    for i in IDX:
        table5.append([handle_name(i) + "x (m)"] +
                      [fmt6(states[t]["points"][i][0]) for t in SHOW_TIMES])
        table5.append([handle_name(i) + "y (m)"] +
                      [fmt6(states[t]["points"][i][1]) for t in SHOW_TIMES])
    table6 = [[handle_name(i) + " (m/s)"] +
              [fmt6(states[t]["speeds"][i]) for t in SHOW_TIMES] for i in IDX]
    paper_sheet("表5 位置", table5, SHOW_TIMES)
    paper_sheet("表6 速度", table6, SHOW_TIMES)

    ws = wb.create_sheet("路径几何")
    ws.append(["项目", "数值"])
    rows = [
        ("盘入螺距 p (m)", f"{PITCH:.6f}"),
        ("调头空间半径 R (m)", f"{TURN_RADIUS:.6f}"),
        ("入口 A 极角 theta_A (rad)", f"{geom['theta_a']:.9f}"),
        ("入口 A (m)", f"({geom['A'][0]:.9f}, {geom['A'][1]:.9f})"),
        ("出口 B=-A (m)", f"({geom['B'][0]:.9f}, {geom['B'][1]:.9f})"),
        ("单位切向 T", f"({geom['T'][0]:.9f}, {geom['T'][1]:.9f})"),
        ("法向 N", f"({geom['N'][0]:.9f}, {geom['N'][1]:.9f})"),
        ("两圆半径和 Q=R1+R2 (m)", f"{geom['Q']:.9f}"),
        ("第一段圆弧半径 R1 (m)", f"{geom['R1']:.9f}"),
        ("第二段圆弧半径 R2 (m)", f"{geom['R2']:.9f}"),
        ("第一段圆心 O1 (m)", f"({geom['O1'][0]:.9f}, {geom['O1'][1]:.9f})"),
        ("第二段圆心 O2 (m)", f"({geom['O2'][0]:.9f}, {geom['O2'][1]:.9f})"),
        ("两圆弧切点 J (m)", f"({geom['J'][0]:.9f}, {geom['J'][1]:.9f})"),
        ("公共圆心角 alpha (rad)", f"{geom['alpha']:.9f}"),
        ("第一段圆弧弧长 l1 (m)", f"{geom['length1']:.9f}"),
        ("第二段圆弧弧长 l2 (m)", f"{geom['length2']:.9f}"),
        ("调头曲线总长 L_turn (m)", f"{geom['turn_length']:.9f}"),
        ("调头路径最大极径 (m)", f"{geom['max_radius_on_turn']:.9f}"),
        ("切点 J 位置连续性误差 (m)", f"{geom['junction_position_error']:.3e}"),
        ("圆心距与外切条件残差 (m)", f"{geom['circle_tangency_error']:.3e}"),
        ("A 点切向夹角 (deg)", f"{geom['tangent_angle_at_A_deg']:.3e}"),
        ("J 点切向夹角 (deg)", f"{geom['tangent_angle_at_J_deg']:.3e}"),
        ("B 点切向夹角 (deg)", f"{geom['tangent_angle_at_B_deg']:.3e}"),
    ]
    for r in rows:
        ws.append(list(r))

    ws = wb.create_sheet("半径比例不变性")
    ws.append(["半径比 lambda=R1/R2", "R1 (m)", "R2 (m)", "调头曲线总长 (m)",
               "调头路径最大极径 (m)", "最大极径是否<=4.5 m"])
    for row in sweep:
        ws.append([
            row["ratio"],
            round(row["R1"], 9),
            round(row["R2"], 9),
            round(row["turn_length"], 9),
            round(row["max_radius_on_turn"], 9),
            "是" if row["max_radius_on_turn"] <= TURN_RADIUS + 1e-12 else "否",
        ])

    ws = wb.create_sheet("说明")
    notes = [
        "2024 高教社杯 A 题 问题4 计算结果 (solve_q4.py 生成)",
        "模型: 盘入螺线 r=b*theta, b=1.7/(2pi) m; 盘出螺线为中心对称镜像 -q(phi);",
        "调头路径为两段外切圆弧, 半径比 R1:R2=2:1, 与两条螺线均相切, 位于半径 4.5 m 圆内;",
        "龙头前把手速率恒为 1 m/s, t=0 位于入口 A; 所有把手沿复合路径满足弦长约束;",
        "速度由弦长约束隐式求导逐节递推; 数值保留 6 位小数。",
        "把手顺序: 龙头前把手(0), 第1~221节龙身前把手(1~221), 龙尾前把手(222), 龙尾后把手(223)。",
    ]
    for line in notes:
        ws.append([line])

    xlsx_path = os.path.join(BASE, "result4.xlsx")
    wb.save(xlsx_path)
    print("saved", xlsx_path)


def write_key_txt(geom: dict, sweep: list, fd: dict, ref_err, chord_resid,
                  states: dict) -> None:
    lines = []
    lines.append("=== 第四问调头路径关键几何结果 (盘入螺距 1.7 m, R=4.5 m, R1:R2=2:1) ===")
    lines.append(f"入口极角 theta_A = {geom['theta_a']:.9f} rad")
    lines.append(f"入口 A = ({geom['A'][0]:.9f}, {geom['A'][1]:.9f}) m, 出口 B = -A = ({geom['B'][0]:.9f}, {geom['B'][1]:.9f}) m")
    lines.append(f"单位切向 T = ({geom['T'][0]:.9f}, {geom['T'][1]:.9f}), 法向 N = ({geom['N'][0]:.9f}, {geom['N'][1]:.9f})")
    lines.append(f"两圆半径和 Q = {geom['Q']:.9f} m")
    lines.append(f"第一段圆弧半径 R1 = {geom['R1']:.9f} m, 第二段圆弧半径 R2 = {geom['R2']:.9f} m")
    lines.append(f"第一段圆心 O1 = ({geom['O1'][0]:.9f}, {geom['O1'][1]:.9f}) m")
    lines.append(f"第二段圆心 O2 = ({geom['O2'][0]:.9f}, {geom['O2'][1]:.9f}) m")
    lines.append(f"两圆弧切点 J = ({geom['J'][0]:.9f}, {geom['J'][1]:.9f}) m")
    lines.append(f"公共圆心角 alpha1 = alpha2 = {geom['alpha']:.9f} rad")
    lines.append(f"第一段弧长 l1 = {geom['length1']:.9f} m, 第二段弧长 l2 = {geom['length2']:.9f} m")
    lines.append(f"调头曲线总长 L_turn = {geom['turn_length']:.9f} m")
    lines.append(f"调头路径最大极径 = {geom['max_radius_on_turn']:.9f} m (<= 4.5 m, 不越出调头空间)")
    lines.append("")
    lines.append("=== 能否调整圆弧使调头曲线变短: 结论 === 不能。")
    lines.append("保持入口 A、出口 B、两端切向与两圆外切不变, 令 R1=lambda*Q, R2=(1-lambda)*Q: ")
    lines.append("Q=||D||^2/[2(D·N)] 与 lambda 无关; 两段圆心角均为 alpha=2arcsin(||D||/(2Q)) 也与 lambda 无关,")
    lines.append("故 L_turn=R1*alpha1+R2*alpha2=Q*alpha 恒定, 改变半径比只能移动切点 J 与重新分配两段长度。")
    lines.append("")
    lines.append("半径比 lambda | R1 (m) | R2 (m) | L_turn (m) | 最大极径 (m)")
    for row in sweep:
        lines.append(f"  {row['ratio']:<12.2f} | {row['R1']:.9f} | {row['R2']:.9f} | "
                     f"{row['turn_length']:.9f} | {row['max_radius_on_turn']:.9f}")
    lines.append("")
    lines.append("=== 数值校验 ===")
    lines.append(f"切点 J 位置连续性误差 = {geom['junction_position_error']:.3e} m")
    lines.append(f"两圆心距与外切条件残差 = {geom['circle_tangency_error']:.3e} m")
    lines.append(f"A 点螺线-圆弧切向夹角 = {geom['tangent_angle_at_A_deg']:.3e} deg")
    lines.append(f"J 点两圆弧切向夹角 = {geom['tangent_angle_at_J_deg']:.3e} deg")
    lines.append(f"B 点圆弧-螺线切向夹角 = {geom['tangent_angle_at_B_deg']:.3e} deg")
    lines.append(f"相邻把手弦长最大残差 (全部201个时刻) = {chord_resid:.3e} m")
    lines.append(f"解析速度与中心有限差分最大相对误差 = {fd['relative']:.3e} (步长 1e-5 s)")
    if ref_err is not None:
        lines.append(f"与工作区独立复现参考结果最大绝对偏差 = {ref_err:.3e} (m 或 m/s)")
    lines.append("")
    lines.append("=== 表5 论文位置表 (单位 m, 保留6位小数) ===")
    lines.append("部位" + "".join("%13s" % ("%d s" % c) for c in SHOW_TIMES))
    for i in IDX:
        lines.append("%-16s" % (handle_name(i) + "x (m)") +
                     "".join("%13.6f" % fmt6(states[t]["points"][i][0]) for t in SHOW_TIMES))
        lines.append("%-16s" % (handle_name(i) + "y (m)") +
                     "".join("%13.6f" % fmt6(states[t]["points"][i][1]) for t in SHOW_TIMES))
    lines.append("")
    lines.append("=== 表6 论文速度表 (单位 m/s, 保留6位小数) ===")
    lines.append("部位" + "".join("%13s" % ("%d s" % c) for c in SHOW_TIMES))
    for i in IDX:
        lines.append("%-16s" % (handle_name(i) + " (m/s)") +
                     "".join("%13.6f" % fmt6(states[t]["speeds"][i]) for t in SHOW_TIMES))
    lines.append("")
    lines.append(f"全量数据 (201个时刻 x 224个把手的位置与速度) 保存在 {os.path.join(BASE, 'result4.xlsx')}")
    txt_path = os.path.join(BASE, "表5_表6_关键结果.txt")
    with open(txt_path, "w", encoding="utf-8") as fp:
        fp.write("\n".join(lines))
    print("\n".join(lines))


def draw_path_diagram(path: TurnPath, geom: dict) -> None:
    fig, ax = plt.subplots(figsize=(9.5, 9.5))
    inward = sample_path(path, -100, 0, 600)
    arc1 = sample_path(path, 0, path.length1, 300)
    arc2 = sample_path(path, path.length1, path.turn_length, 300)
    outward = sample_path(path, path.turn_length, path.turn_length + 100, 600)
    ax.plot(inward[:, 0], inward[:, 1], "-", color="#2c7fb8", linewidth=1.4,
            label="盘入螺线 (s<0)")
    ax.plot(arc1[:, 0], arc1[:, 1], "-", color="#c0392b", linewidth=2.2,
            label="第一段圆弧 R1")
    ax.plot(arc2[:, 0], arc2[:, 1], "-", color="#e67e22", linewidth=2.2,
            label="第二段圆弧 R2")
    ax.plot(outward[:, 0], outward[:, 1], "-", color="#27ae60", linewidth=1.4,
            label="盘出螺线 (s>L)")
    ax.add_patch(Circle((0, 0), TURN_RADIUS, fill=False, color="#8e44ad",
                        linestyle="--", linewidth=1.6, label="调头空间边界 R=4.5 m"))
    ax.plot(*path.a, "o", color="#c0392b", markersize=8, label="入口 A")
    ax.plot(*path.c, "s", color="#1f4e79", markersize=9, label="切点 J")
    ax.plot(-path.a[0], -path.a[1], "o", color="#27ae60", markersize=8, label="出口 B")
    ax.plot(*path.o1, "x", color="#c0392b", markersize=9, label="圆心 O1")
    ax.plot(*path.o2, "x", color="#e67e22", markersize=9, label="圆心 O2")
    ax.plot([path.o1[0], path.a[0]], [path.o1[1], path.a[1]], "-", color="#999999",
            linewidth=0.7)
    ax.plot([path.o1[0], path.c[0]], [path.o1[1], path.c[1]], "-", color="#999999",
            linewidth=0.7)
    ax.plot([path.o2[0], path.c[0]], [path.o2[1], path.c[1]], "-", color="#999999",
            linewidth=0.7)
    ax.plot([path.o2[0], -path.a[0]], [path.o2[1], -path.a[1]], "-", color="#999999",
            linewidth=0.7)
    ax.set_aspect("equal", adjustable="box")
    ax.grid(True, linestyle=":", linewidth=0.6, alpha=0.6)
    ax.set_title("第四问 盘入螺线-S形双圆弧-盘出螺线调头路径")
    ax.set_xlabel("x (m)")
    ax.set_ylabel("y (m)")
    ax.legend(loc="best", fontsize=9)
    fig.tight_layout()
    png = os.path.join(BASE, "调头路径示意图.png")
    fig.savefig(png, dpi=170)
    plt.close(fig)
    print("saved", png)


def draw_snapshots(path: TurnPath, states: dict) -> None:
    fig, ax = plt.subplots(figsize=(9.5, 9.5))
    inward = sample_path(path, -100, 0, 600)
    arc1 = sample_path(path, 0, path.length1, 200)
    arc2 = sample_path(path, path.length1, path.turn_length, 200)
    outward = sample_path(path, path.turn_length, path.turn_length + 100, 600)
    ax.plot(inward[:, 0], inward[:, 1], "-", color="#bdc3c7", linewidth=0.9,
            label="复合路径参考线")
    ax.plot(arc1[:, 0], arc1[:, 1], "-", color="#bdc3c7", linewidth=0.9)
    ax.plot(arc2[:, 0], arc2[:, 1], "-", color="#bdc3c7", linewidth=0.9)
    ax.plot(outward[:, 0], outward[:, 1], "-", color="#bdc3c7", linewidth=0.9)
    colors = {-100: "#2c7fb8", -50: "#27ae60", 0: "#c0392b", 50: "#e67e22", 100: "#8e44ad"}
    for t in SHOW_TIMES:
        points = states[t]["points"]
        ax.plot(points[:, 0], points[:, 1], "-", color=colors[t], linewidth=1.0,
                label=f"t = {t:+d} s 全龙")
        ax.plot(points[0, 0], points[0, 1], "o", color=colors[t], markersize=6)
    ax.add_patch(Circle((0, 0), TURN_RADIUS, fill=False, color="#34495e",
                        linestyle="--", linewidth=1.4, label="调头空间边界 R=4.5 m"))
    ax.set_aspect("equal", adjustable="box")
    ax.grid(True, linestyle=":", linewidth=0.6, alpha=0.6)
    ax.set_title("第四问 调头全过程 (-100 s ~ 100 s) 全龙形态")
    ax.set_xlabel("x (m)")
    ax.set_ylabel("y (m)")
    ax.legend(loc="best", fontsize=9)
    fig.tight_layout()
    png = os.path.join(BASE, "调头全过程形态图.png")
    fig.savefig(png, dpi=170)
    plt.close(fig)
    print("saved", png)


def main() -> None:
    path = TurnPath(PITCH, TURN_RADIUS, RATIO)
    geom = geometry_checks(path)
    sweep = invariance_sweep()
    states = compute_states(path)
    fd = finite_difference_check(path, states)
    chord_resid = chord_residual_max(states)
    ref_err = reference_comparison(states)

    write_result4(path, geom, states, sweep, fd, ref_err)
    write_key_txt(geom, sweep, fd, ref_err, chord_resid, states)
    draw_path_diagram(path, geom)
    draw_snapshots(path, states)

    print("A =", geom["A"])
    print("B =", geom["B"])
    print("J =", geom["J"])
    print("R1 = %.9f, R2 = %.9f, L_turn = %.9f" % (geom["R1"], geom["R2"], geom["turn_length"]))
    print("max radius on turn = %.12f" % geom["max_radius_on_turn"])
    print("tangent angles (deg): A=%.3e J=%.3e B=%.3e" % (
        geom["tangent_angle_at_A_deg"], geom["tangent_angle_at_J_deg"],
        geom["tangent_angle_at_B_deg"]))
    print("chord residual max = %.3e, FD rel = %.3e, ref err = %s" % (
        chord_resid, fd["relative"], ref_err))


if __name__ == "__main__":
    main()

