# -*- coding: utf-8 -*-
"""
2024 高教社杯全国大学生数学建模竞赛 A 题 问题 1 独立求解程序
=================================================================
模型要点
--------
1) 等距螺线(阿基米德螺线): 极坐标 r = a*theta,  a = p/(2*pi),  p = 0.55 m 为螺距。
   直角坐标: P(theta) = (a*theta*cos(theta), a*theta*sin(theta)).
2) 顺时针盘入: theta 随时间减小; 初始 theta0 = 32*pi (第 16 圈末端), A = (8.8, 0) m.
3) 龙头前把手速率恒为 1 m/s, 即弧长 s(theta0(t)) = s(32*pi) - t, 用 Newton 法反解 theta.
4) 所有把手中心位于螺线上, 相邻把手中心距 = 板凳两端孔心距:
   龙头板 2.86 m, 其余板 1.65 m. 龙身沿龙头走过的路径向螺线外侧延伸,
   故对每一节用"弦长精确求解": |P(theta_{i+1}) - P(theta_i)| = L_i 且 theta_{i+1} > theta_i.
5) 速度由弦长约束 |P_i - P_{i-1}| = L_i 对时间隐式求导逐节递推:
   (P_i-P_{i-1})/L . V_i = (P_i-P_{i-1})/L . V_{i-1},  V_i = v_i * e_i,
   e_i 为该把手沿螺线行进方向的单位切向量(顺时针, 即 -P'(theta)/|P'(theta)|),
   得 v_i = v_{i-1} * (link.e_{i-1}) / (link.e_i).
输出
----
result1.xlsx : 全表 0~300 s 每秒 224 个把手的位置(Sheet 位置)与速度(Sheet 速度),
               以及论文表 1 / 表 2(Sheet 表1 位置 / 表2 速度), 数值保留 6 位小数。
表1_表2_关键结果.txt : 论文表 1、表 2 与数值校验结果。
"""
import json
import math
import os
from openpyxl import Workbook

# ---------------- 常量 ----------------
PITCH   = 0.55                        # 螺距 p (m)
A       = PITCH / (2.0 * math.pi)     # 螺线系数 a = p/(2pi) (m/rad)
THETA0  = 32.0 * math.pi              # 初始龙头极角 (第16圈末端, +x 轴)
HEAD_V  = 1.0                         # 龙头前把手速率 (m/s)
L_HEAD  = 3.41 - 2 * 0.275            # 龙头板两孔心距 = 2.86 m
L_BODY  = 2.20 - 2 * 0.275            # 龙身/龙尾板两孔心距 = 1.65 m
N       = 224                         # 把手总数 (223 节板凳)
T_END   = 300

HERE = os.path.dirname(os.path.abspath(__file__))

# ---------------- 螺线基本函数 ----------------
def point(th):
    """P(theta) = (a*theta*cos(theta), a*theta*sin(theta))"""
    c, s = math.cos(th), math.sin(th)
    return (A * th * c, A * th * s)

def deriv(th):
    """P'(theta) = dP/dtheta = a*(cos-th*sin, sin+th*cos)"""
    c, s = math.cos(th), math.sin(th)
    return (A * (c - th * s), A * (s + th * c))

def arc_s(th):
    """螺线从极点 (theta=0) 到 theta 的弧长 s(theta) = (a/2)[theta*sqrt(theta^2+1) + asinh(theta)]"""
    return 0.5 * A * (th * math.sqrt(th * th + 1.0) + math.asinh(th))

def theta_from_arc(s):
    """由弧长 s 反解 theta: Newton 迭代, f = s(theta)-s, f' = a*sqrt(theta^2+1)"""
    if s <= 0.0:
        return 0.0
    th = math.sqrt(2.0 * s / A)          # 大 theta 渐近: s ~ a*theta^2/2
    for _ in range(80):
        f = arc_s(th) - s
        if abs(f) < 1e-15:
            break
        df = A * math.sqrt(th * th + 1.0)
        th -= f / df
    return th

def link_len(i):
    """把手 i-1 与把手 i 之间的板长(孔心距): i=1 为龙头板 2.86 m, 其余 1.65 m"""
    return L_HEAD if i == 1 else L_BODY

def chord_next(theta_prev, L):
    """已知前一把手极角 theta_prev, 求下一把手极角 theta_next > theta_prev,
    使弦长 |P(theta_next)-P(theta_prev)| = L (Newton 迭代, 必要时二分兜底)。"""
    px, py = point(theta_prev)
    th = theta_prev + L / math.hypot(A * theta_prev, A)   # 初始: 以弧长间距近似 (弦 < 弧)
    for _ in range(100):
        x, y = point(th)
        dx, dy = x - px, y - py
        f = dx * dx + dy * dy - L * L
        if abs(f) < 1e-24:
            break
        c, s = math.cos(th), math.sin(th)
        gx, gy = A * (c - th * s), A * (s + th * c)
        df = 2.0 * (dx * gx + dy * gy)
        th_new = th - f / df
        if th_new <= theta_prev or th_new - theta_prev > 0.75 * math.pi:
            # Newton 越界 -> 二分 (区间内弦长随极角差单调增)
            lo, hi = theta_prev, theta_prev + 0.75 * math.pi
            for _ in range(100):
                mid = 0.5 * (lo + hi)
                mx, my = point(mid)
                fmid = (mx - px) ** 2 + (my - py) ** 2 - L * L
                if fmid < 0.0:
                    lo = mid
                else:
                    hi = mid
                if hi - lo < 1e-16:
                    th = mid
                    break
            break
        th = th_new
    return th

def build_chain(theta_head):
    """由龙头极角构建整条龙: 224 个把手均落在螺线上, 相邻弦长 = 板长"""
    ths = [theta_head]
    for i in range(1, N):
        ths.append(chord_next(ths[-1], link_len(i)))
    return ths

def speeds(ths):
    """刚性弦长约束的速度递推: v_i = v_{i-1} * (link.e_{i-1}) / (link.e_i)"""
    v = [HEAD_V]
    for i in range(1, N):
        x0, y0 = point(ths[i - 1])
        x1, y1 = point(ths[i])
        lx, ly = x1 - x0, y1 - y0
        ln = math.hypot(lx, ly)
        lx, ly = lx / ln, ly / ln
        d0x, d0y = deriv(ths[i - 1])
        n0 = math.hypot(d0x, d0y)
        d1x, d1y = deriv(ths[i])
        n1 = math.hypot(d1x, d1y)
        e0 = (-d0x / n0, -d0y / n0)      # 顺时针行进的单位切向量
        e1 = (-d1x / n1, -d1y / n1)
        num = lx * e0[0] + ly * e0[1]
        den = lx * e1[0] + ly * e1[1]
        v.append(v[-1] * num / den)
    return v

# ---------------- 主计算: 0~300 s 每秒 ----------------
S0 = arc_s(THETA0)
times = list(range(T_END + 1))
POS = []          # POS[t][i] = (x, y)
VEL = []
THS = []          # THS[t][i] = theta
for t in times:
    th_head = theta_from_arc(S0 - HEAD_V * t)
    ths = build_chain(th_head)
    THS.append(ths)
    POS.append([point(th) for th in ths])
    VEL.append(speeds(ths))

# ---------------- 数值校验 ----------------
def dist(p0, p1):
    return math.hypot(p1[0] - p0[0], p1[1] - p0[1])

chord_resid = 0.0
for t in times:
    for i in range(1, N):
        chord_resid = max(chord_resid, abs(dist(POS[t][i - 1], POS[t][i]) - link_len(i)))

arc_resid = max(abs((arc_s(THS[t][0]) - (S0 - HEAD_V * t))) for t in times)

# 速度有限差分校核 (t=100, 对 theta0 作中心差分, 经链映射传递)
tp = 100
h = 1e-5
thA = build_chain(theta_from_arc(S0 - HEAD_V * (tp - h)))   # 扰动 theta0 的等效链
thB = build_chain(theta_from_arc(S0 - HEAD_V * (tp + h)))
thC = THS[tp]
fd_rel = 0.0
for i in range(N):
    th_fd = (thA[i] - thB[i]) / (2.0 * h)                    # dtheta_i/dt 中心差分
    v_fd = abs(th_fd) * math.hypot(A * thC[i], A)
    fd_rel = max(fd_rel, abs(v_fd - VEL[tp][i]) / VEL[tp][i])

# 简化"弧长间距模型"(相邻把手沿螺线相距 L 的弧长)与弦长模型的差距
th_arc = [THETA0]
for i in range(1, N):
    th_arc.append(theta_from_arc(arc_s(th_arc[-1]) + link_len(i)))
diff_arc_model = max(dist(point(THS[0][i]), point(th_arc[i])) for i in range(N))
th_arc300 = [THS[300][0]]
for i in range(1, N):
    th_arc300.append(theta_from_arc(arc_s(th_arc300[-1]) + link_len(i)))
diff_arc_model_300 = max(dist(point(THS[300][i]), point(th_arc300[i])) for i in range(N))

# ---------------- 论文表 1 / 表 2 的取样 ----------------
IDX = [0, 1, 51, 101, 151, 201, 223]         # 龙头前; 第1/51/101/151/201节龙身前; 龙尾后
TT  = [0, 60, 120, 180, 240, 300]

def handle_name(i):
    if i == 0:
        return "龙头"
    if 1 <= i <= 221:
        return "第%d节龙身" % i
    if i == 222:
        return "龙尾"
    return "龙尾（后）"

# 与工作区独立复现结果对照 (仅校验, 本程序为独立实现)
ref = {}
ref_path = os.path.join(HERE, "..", "24A题_独立复现", "independent_results.json")
if os.path.exists(ref_path):
    with open(ref_path, encoding="utf-8") as fp:
        ref = json.load(fp).get("problem1", {})
ref_err = None
if ref:
    errs = []
    for key, val in ref.items():
        t = int(key)
        if 0 <= t <= T_END:
            posr = val["positions"]     # 参考 json 仅存 7 个论文关键把手
            sprr = val["speeds"]
            for j, i in enumerate(IDX):
                errs.append(max(abs(POS[t][i][0] - posr[j][0]), abs(POS[t][i][1] - posr[j][1]),
                                abs(VEL[t][i] - sprr[j])))
    ref_err = max(errs)



table1 = []   # 行: 14 条 (7 把手 x/y), 列: 标签 + 6 时刻
for i in IDX:
    table1.append([handle_name(i) + "x (m)"] + [round(POS[t][i][0], 6) for t in TT])
    table1.append([handle_name(i) + "y (m)"] + [round(POS[t][i][1], 6) for t in TT])
table2 = [[handle_name(i) + " (m/s)"] + [round(VEL[t][i], 6) for t in TT] for i in IDX]

# ---------------- 输出 result1.xlsx ----------------
wb = Workbook()

def sheet_positions(wb):
    ws = wb.active
    ws.title = "位置"
    header = ["t (s)"]
    for i in range(N):
        header += [handle_name(i) + "x (m)", handle_name(i) + "y (m)"]
    ws.append(header)
    for t in times:
        row = [t]
        for i in range(N):
            row += [round(POS[t][i][0], 6), round(POS[t][i][1], 6)]
        ws.append(row)
    return ws

def sheet_speeds(wb):
    ws = wb.create_sheet("速度")
    header = ["t (s)"] + [handle_name(i) + " (m/s)" for i in range(N)]
    ws.append(header)
    for t in times:
        ws.append([t] + [round(VEL[t][i], 6) for i in range(N)])
    return ws

def sheet_paper(wb, title, rows, cols):
    ws = wb.create_sheet(title)
    ws.append([" "] + ["%d s" % c for c in cols])
    for r in rows:
        ws.append(r)
    return ws

sheet_positions(wb)
sheet_speeds(wb)
sheet_paper(wb, "表1 位置", table1, TT)
sheet_paper(wb, "表2 速度", table2, TT)

ws = wb.create_sheet("说明")
notes = [
    "2024 高教社杯 A 题 问题1 计算结果 (solve_q1.py 生成)",
    "模型: 等距螺线 r=a*theta, a=0.55/(2pi) m; 龙头前把手速率 1 m/s 顺时针盘入;",
    "所有把手中心位于螺线上, 相邻把手弦长精确等于板长 (龙头板 2.86 m, 其余 1.65 m);",
    "速度由弦长约束隐式求导逐节递推; 数值保留 6 位小数。",
    "把手顺序: 龙头前把手(0), 第1~221节龙身前把手(1~221), 龙尾前把手(222), 龙尾后把手(223)。",
]
for line in notes:
    ws.append([line])

xlsx_path = os.path.join(HERE, "result1.xlsx")
wb.save(xlsx_path)

# ---------------- 输出关键结果 txt ----------------
lines = []
lines.append("=== 表1 论文位置表 (单位 m, 保留6位小数) ===")
lines.append("部位" + "".join("%12s" % ("%d s" % c) for c in TT))
for r in table1:
    lines.append("%-14s" % r[0] + "".join("%12.6f" % v for v in r[1:]))
lines.append("")
lines.append("=== 表2 论文速度表 (单位 m/s, 保留6位小数) ===")
lines.append("部位" + "".join("%12s" % ("%d s" % c) for c in TT))
for r in table2:
    lines.append("%-14s" % r[0] + "".join("%12.6f" % v for v in r[1:]))
lines.append("")
lines.append("=== 数值校验 ===")
lines.append("龙头初始位置: (%.9f, %.9f) m" % POS[0][0])
lines.append("龙头 t=60s 位置: (%.9f, %.9f) m" % POS[60][0])
lines.append("相邻把手弦长最大残差: %.3e m (板长约束满足程度)" % chord_resid)
lines.append("龙头 0->300s 弧长减少: %.9f m (理论 300.000000)" % (S0 - arc_s(THS[300][0])))
lines.append("速度递推与中心有限差分最大相对误差: %.3e" % fd_rel)
lines.append("弦长模型与弧长间距模型位置最大偏差(t=0): %.6f m; (t=300): %.6f m" % (diff_arc_model, diff_arc_model_300))
lines.append("龙头板(2.86m)相邻把手速比示例 t=0: %.9f" % (VEL[0][1] / VEL[0][0]))
if ref_err is not None:
    lines.append("与独立复现参考结果最大绝对偏差: %.3e (m 或 m/s)" % ref_err)
lines.append("")
lines.append("输出文件: %s" % xlsx_path)
txt_path = os.path.join(HERE, "表1_表2_关键结果.txt")
with open(txt_path, "w", encoding="utf-8") as fp:
    fp.write("\n".join(lines))
print("\n".join(lines))




