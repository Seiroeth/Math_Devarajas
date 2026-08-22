# -*- coding: utf-8 -*-
"""2024 CUMCM Problem A 第二问求解脚本。

输出:
  result2.xlsx            终止时刻全部 224 个把手的位置与速度
  表3_终止时刻关键结果.txt  论文要求的关键把手位置/速度表
  终止时刻_碰撞示意图.png    终止时刻整条龙的构型与首次接触板凳对
"""
import os
import math

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import Polygon
from openpyxl import Workbook

from common_model import (
    N_BOARDS,
    N_HANDLES,
    BOARD_WIDTH,
    END_OVERHANG,
    handle_distance,
    selected_handle_indices,
    find_terminal_time,
    head_theta_at_time,
    spiral_b,
    spiral_point,
    spiral_forward_tangent,
    constraint_residual,
)

plt.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei"]
plt.rcParams["axes.unicode_minus"] = False

BASE = os.path.dirname(os.path.abspath(__file__))


def handle_label(i: int) -> str:
    """把手序号 i 的论文中文名称，i=0..223。"""
    if i == 0:
        return "龙头前把手"
    if 1 <= i <= 221:
        return f"第{i}节龙身前把手"
    if i == 222:
        return "龙尾前把手"
    return "龙尾后把手"


def board_label(k: int) -> str:
    """板凳编号 k=1..223。"""
    if k == 1:
        return "第1节（龙头）"
    if 2 <= k <= 222:
        return f"第{k}节（第{k-1}节龙身）"
    return "第223节（龙尾）"


def board_rectangle(points: np.ndarray, k: int):
    """返回第 k 节板凳（1-based）的矩形角点，用于碰撞示意图。"""
    a = points[k - 1]
    b = points[k]
    d = handle_distance(k)
    axis_long = (b - a) / d
    axis_wide = np.array([-axis_long[1], axis_long[0]])
    center = 0.5 * (a + b)
    half_length = 0.5 * (d + 2 * END_OVERHANG)
    half_width = 0.5 * BOARD_WIDTH
    corners = np.array([
        center - half_length * axis_long - half_width * axis_wide,
        center + half_length * axis_long - half_width * axis_wide,
        center + half_length * axis_long + half_width * axis_wide,
        center - half_length * axis_long + half_width * axis_wide,
    ])
    return corners


def write_result2(terminal: dict) -> None:
    points = terminal["points"]
    speeds = np.abs(terminal["speeds"])

    wb = Workbook()

    ws_pos = wb.active
    ws_pos.title = "位置"
    ws_pos.append(["把手序号", "部位", "x(m)", "y(m)"])
    for i in range(N_HANDLES):
        ws_pos.append([i, handle_label(i), round(float(points[i, 0]), 6),
                       round(float(points[i, 1]), 6)])

    ws_vel = wb.create_sheet("速度")
    ws_vel.append(["把手序号", "部位", "vx(m/s)", "vy(m/s)", "速率(m/s)"])
    tangents = None
    for i in range(N_HANDLES):
        # 计算单位切向速度向量，速率来自弦长约束递推。
        if i == 0:
            theta = head_theta_at_time(terminal["time"], 0.55, 1.0)
            tangent = np.asarray(spiral_forward_tangent(theta, 0.55), dtype=float)
        else:
            # 前一个把手沿螺线更靠内，当前把手的瞬时前进方向近似指向前一个把手。
            delta = points[i - 1] - points[i]
            tangent = delta / (np.linalg.norm(delta) + 1e-15)
        vx = float(speeds[i] * tangent[0])
        vy = float(speeds[i] * tangent[1])
        ws_vel.append([i, handle_label(i), round(vx, 6), round(vy, 6),
                       round(float(speeds[i]), 6)])

    ws_key = wb.create_sheet("关键结果")
    ws_key.append(["项目", "数值"])
    ws_key.append(["终止时刻 t* (s)", f"{terminal['time']:.9f}"])
    ws_key.append(["龙头极角 theta* (rad)", f"{terminal['theta_head']:.9f}"])
    r_head = float(np.linalg.norm(points[0]))
    ws_key.append(["龙头极径 r* (m)", f"{r_head:.9f}"])
    ws_key.append(["首次接触板凳对", f"第{terminal['pair'][0]}节 与 第{terminal['pair'][1]}节"])
    ws_key.append(["接触时刻最小间隙 (m)", f"{terminal['clearance']:.6e}"])
    ws_key.append(["相邻把手弦长最大残差 (m)", f"{terminal['distance_residual']:.6e}"])

    xlsx_path = os.path.join(BASE, "result2.xlsx")
    wb.save(xlsx_path)
    print("saved", xlsx_path)

    selected = selected_handle_indices()
    rows = []
    for idx in selected:
        rows.append([
            handle_label(idx),
            f"{float(points[idx, 0]):.6f}",
            f"{float(points[idx, 1]):.6f}",
            f"{float(np.abs(speeds[idx])):.6f}",
        ])
    header = ["部位", "x(m)", "y(m)", "速度(m/s)"]
    with open(os.path.join(BASE, "表3_终止时刻关键结果.txt"), "w", encoding="utf-8") as f:
        f.write(f"=== 第二问终止时刻关键结果（t* = {terminal['time']:.9f} s）===\n")
        f.write("首次接触板凳对: 第%d节 与 第%d节\n" % tuple(terminal["pair"]))
        f.write("龙头极角: %.9f rad, 龙头极径: %.9f m\n" % (terminal["theta_head"], r_head))
        f.write("最小间隙: %.6e m\n\n" % terminal["clearance"])
        f.write("\t".join(header) + "\n")
        for row in rows:
            f.write("\t".join(row) + "\n")
    print("saved", os.path.join(BASE, "表3_终止时刻关键结果.txt"))


def draw_figure(terminal: dict) -> None:
    points = terminal["points"]
    pair = terminal["pair"]

    fig, ax = plt.subplots(figsize=(9.5, 9.5))
    for k in range(1, N_BOARDS + 1):
        corners = board_rectangle(points, k)
        if k in pair:
            color = "#e74c3c"
            alpha = 0.75
            zorder = 5
        else:
            color = "#2c7fb8"
            alpha = 0.14
            zorder = 1
        ax.add_patch(Polygon(corners, closed=True, fill=True, facecolor=color,
                             edgecolor="none", alpha=alpha, zorder=zorder))
    ax.plot(points[:, 0], points[:, 1], "-", color="#1f4e79", linewidth=0.9,
            label="把手中心连线", zorder=3)
    ax.plot(points[0, 0], points[0, 1], "o", color="#c0392b", markersize=7,
            label="龙头前把手", zorder=6)
    ax.plot(points[pair[0] - 1, 0], points[pair[0] - 1, 1], "s", color="#e67e22",
            markersize=9, label=f"碰撞板 {pair[0]} 前端")
    ax.plot(points[pair[0], 0], points[pair[0], 1], "s", color="#e67e22",
            markersize=9)
    ax.plot(points[pair[1] - 1, 0], points[pair[1] - 1, 1], "^", color="#8e44ad",
            markersize=9, label=f"碰撞板 {pair[1]} 前端")
    ax.plot(points[pair[1], 0], points[pair[1], 1], "^", color="#8e44ad",
            markersize=9)

    ax.set_aspect("equal", adjustable="box")
    ax.grid(True, linestyle=":", linewidth=0.6, alpha=0.6)
    ax.set_title(f"第二问终止时刻构型（t* = {terminal['time']:.6f} s）")
    ax.set_xlabel("x (m)")
    ax.set_ylabel("y (m)")
    ax.legend(loc="best", fontsize=9)
    fig.tight_layout()
    png = os.path.join(BASE, "终止时刻_碰撞示意图.png")
    fig.savefig(png, dpi=170)
    plt.close(fig)
    print("saved", png)


def main() -> None:
    terminal = find_terminal_time()
    terminal["distance_residual"] = constraint_residual(terminal["points"])
    write_result2(terminal)
    draw_figure(terminal)
    print("t* =", f"{terminal['time']:.9f}")
    print("pair =", terminal["pair"])
    print("head theta =", f"{terminal['theta_head']:.9f}")
    print("distance residual =", f"{terminal['distance_residual']:.3e}")


if __name__ == "__main__":
    main()
