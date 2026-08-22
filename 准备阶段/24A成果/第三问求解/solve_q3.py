# -*- coding: utf-8 -*-
"""2024 CUMCM Problem A 第三问求解脚本。

问题: 从盘入到盘出，舞龙队需在直径为 9 m 的圆形调头空间内调头。
求最小螺距 p_min，使龙头前把手能沿相应等距螺线盘入到调头空间边界
(半径 R = 4.5 m)，且从初始位置到边界的盘入全过程 223 块板凳互不碰撞。

方法: 对候选螺距 p，由边界条件 r = b*theta = R 得 theta(p) = R/b，
用问题 1 的弦长递推构建全龙构型，用问题 2 的矩形分离轴(SAT)模型计算
全过程最小安全裕量 H(p) = min_{r0 >= R} g(p, r0)，
再对可行域 H(p) >= 0 做外层二分，得到 p_min。

输出:
  result3.xlsx                最小螺距关键结果、二分收敛过程、边界/临界构型数据
  表4_最小螺距关键结果.txt     论文关键结果与指定把手表
  螺距_安全裕量曲线.png        H(p) 随螺距变化曲线与二分定位
  临界构型_碰撞示意图.png      全过程最小裕量处的全龙构型，标注临界板凳对
"""
import os
import math

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import Polygon, Circle
from openpyxl import Workbook

from common_model import (
    N_BOARDS,
    N_HANDLES,
    BOARD_WIDTH,
    END_OVERHANG,
    INITIAL_THETA,
    handle_distance,
    selected_handle_indices,
    spiral_b,
    spiral_point,
    spiral_forward_tangent,
    spiral_state,
    minimum_clearance,
    build_spiral_chain,
    boundary_clearance_for_pitch,
    path_clearance_for_pitch,
    find_minimum_pitch,
    constraint_residual,
)

plt.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei"]
plt.rcParams["axes.unicode_minus"] = False

BASE = os.path.dirname(os.path.abspath(__file__))
TURN_RADIUS = 4.5


def handle_label(i: int) -> str:
    """把手序号 i 的论文中文名称，i=0..223。"""
    if i == 0:
        return "龙头前把手"
    if 1 <= i <= 221:
        return f"第{i}节龙身前把手"
    if i == 222:
        return "龙尾前把手"
    return "龙尾后把手"


def board_label(k: int) -> str:
    """板凳编号 k=1..223。"""
    if k == 1:
        return "第1节（龙头）"
    if 2 <= k <= 222:
        return f"第{k}节（第{k-1}节龙身）"
    return "第223节（龙尾）"


def board_rectangle(points: np.ndarray, k: int) -> np.ndarray:
    """返回第 k 节板凳（1-based）的矩形角点。"""
    a = points[k - 1]
    b = points[k]
    d = handle_distance(k)
    axis_long = (b - a) / d
    axis_wide = np.array([-axis_long[1], axis_long[0]])
    center = 0.5 * (a + b)
    half_length = 0.5 * (d + 2 * END_OVERHANG)
    half_width = 0.5 * BOARD_WIDTH
    corners = np.array([
        center - half_length * axis_long - half_width * axis_wide,
        center + half_length * axis_long - half_width * axis_wide,
        center + half_length * axis_long + half_width * axis_wide,
        center - half_length * axis_long + half_width * axis_wide,
    ])
    return corners


def bisect_minimum_pitch(low: float, high: float, tol: float = 1e-9,
                         max_iter: int = 40, grid_size: int = 31) -> dict:
    """外层二分求最小螺距，并记录每次迭代的区间与安全裕量。"""
    history = []
    low_state = path_clearance_for_pitch(low, grid_size=grid_size)
    high_state = path_clearance_for_pitch(high, grid_size=grid_size)
    if not (low_state["clearance"] < 0 <= high_state["clearance"]):
        raise RuntimeError(
            f"二分端点无效: H({low:.6f})={low_state['clearance']:.6e}, "
            f"H({high:.6f})={high_state['clearance']:.6e}"
        )
    for k in range(1, max_iter + 1):
        mid = 0.5 * (low + high)
        state = path_clearance_for_pitch(mid, grid_size=grid_size)
        margin = state["clearance"]
        history.append({
            "iter": k,
            "low": low,
            "high": high,
            "mid": mid,
            "margin": margin,
        })
        if margin >= 0.0:
            high = mid
        else:
            low = mid
        if high - low <= tol:
            break
    return {"pitch": high, "history": history}


def write_result3(result: dict) -> None:
    pitch = result["pitch"]
    boundary = result["boundary"]
    path_min = result["path_min"]
    critical = result["critical"]
    verification = result["verification"]
    history = result["history"]

    wb = Workbook()

    ws_key = wb.active
    ws_key.title = "关键结果"
    ws_key.append(["项目", "数值"])
    key_rows = [
        ("最小螺距 p_min (m)", f"{pitch:.9f}"),
        ("调头空间半径 R (m)", f"{TURN_RADIUS:.6f}"),
        ("边界条件龙头极角 theta(p_min) (rad)", f"{boundary['theta']:.9f}"),
        ("边界条件龙头极径 (m)", f"{TURN_RADIUS:.9f}"),
        ("全过程最小安全裕量 H(p_min) (m)", f"{path_min['clearance']:.6e}"),
        ("临界接触板凳对", f"第{path_min['pair'][0]}节 与 第{path_min['pair'][1]}节"),
        ("临界龙头极角 (rad)", f"{path_min['theta']:.9f}"),
        ("临界龙头极径 (m)", f"{path_min['theta'] * spiral_b(pitch):.9f}"),
        ("边界构型安全裕量 (m)", f"{boundary['clearance']:.6e}"),
        ("边界构型最紧板凳对", f"第{boundary['pair'][0]}节 与 第{boundary['pair'][1]}节"),
        ("验证 H(p_min-0.001) (m)", f"{verification['below']:.6e}"),
        ("验证 H(p_min+0.001) (m)", f"{verification['above']:.6e}"),
        ("相邻把手弦长最大残差 (m)", f"{result['residual']:.6e}"),
    ]
    for row in key_rows:
        ws_key.append(list(row))

    ws_hist = wb.create_sheet("二分收敛过程")
    ws_hist.append(["迭代", "p_L (m)", "p_U (m)", "p_M (m)", "H(p_M) (m)"])
    for h in history:
        ws_hist.append([h["iter"], round(h["low"], 12), round(h["high"], 12),
                        round(h["mid"], 12), h["margin"]])

    for sheet_name, cfg in (
        ("边界构型位置", boundary["positions"]),
        ("临界构型位置", critical["positions"]),
    ):
        ws = wb.create_sheet(sheet_name)
        ws.append(["把手序号", "部位", "x(m)", "y(m)"])
        for i in range(N_HANDLES):
            ws.append([i, handle_label(i), round(float(cfg[i, 0]), 6),
                       round(float(cfg[i, 1]), 6)])

    for sheet_name, cfg in (
        ("边界构型速度", boundary),
        ("临界构型速度", critical),
    ):
        ws = wb.create_sheet(sheet_name)
        ws.append(["把手序号", "部位", "vx(m/s)", "vy(m/s)", "速率(m/s)"])
        for i in range(N_HANDLES):
            tangent = cfg["tangents"][i]
            speed = float(abs(cfg["speeds"][i]))
            ws.append([i, handle_label(i), round(float(speed * tangent[0]), 6),
                       round(float(speed * tangent[1]), 6), round(speed, 6)])

    ws_note = wb.create_sheet("说明")
    notes = [
        "2024 高教社杯 A 题 问题3 计算结果 (solve_q3.py 生成)",
        "模型: 等距螺线 r=b*theta, b=p/(2pi); 把手中心位于螺线上, 相邻把手弦长精确等于板长;",
        "碰撞按带宽度0.30 m、端部伸出0.275 m的矩形(分离轴定理)判定, 相邻板凳不计;",
        "最小螺距定义为 H(p)=min_{r0>=4.5m} g(p,r0) >= 0 的最小 p, 外层二分求根;",
        "把手顺序: 龙头前把手(0), 第1~221节龙身前把手(1~221), 龙尾前把手(222), 龙尾后把手(223)。",
    ]
    for line in notes:
        ws_note.append([line])

    xlsx_path = os.path.join(BASE, "result3.xlsx")
    wb.save(xlsx_path)
    print("saved", xlsx_path)


def write_key_txt(result: dict) -> None:
    pitch = result["pitch"]
    boundary = result["boundary"]
    path_min = result["path_min"]
    critical = result["critical"]
    verification = result["verification"]

    lines = []
    lines.append("=== 第三问最小螺距关键结果 ===")
    lines.append(f"最小螺距 p_min = {pitch:.9f} m")
    lines.append(f"调头空间半径 R = {TURN_RADIUS:.6f} m")
    lines.append(f"边界条件龙头极角 theta(p_min) = {boundary['theta']:.9f} rad")
    lines.append(f"全过程最小安全裕量 H(p_min) = {path_min['clearance']:.6e} m (≈0, 临界接触)")
    lines.append(f"临界接触板凳对: 第{path_min['pair'][0]}节 与 第{path_min['pair'][1]}节")
    lines.append(f"临界龙头极角 = {path_min['theta']:.9f} rad, "
                 f"临界龙头极径 = {path_min['theta'] * spiral_b(pitch):.9f} m")
    lines.append(f"边界构型安全裕量 = {boundary['clearance']:.6e} m (>0, 到达边界时仍安全)")
    lines.append(f"验证: H(p_min-0.001) = {verification['below']:.6e} m < 0 不可行; "
                 f"H(p_min+0.001) = {verification['above']:.6e} m > 0 可行")
    lines.append("")
    lines.append("--- 临界构型指定把手位置与速度 ---")
    lines.append("部位\tx(m)\ty(m)\t速度(m/s)")
    selected = selected_handle_indices()
    for idx in selected:
        p = critical["positions"][idx]
        lines.append(f"{handle_label(idx)}\t{p[0]:.6f}\t{p[1]:.6f}"
                     f"\t{abs(critical['speeds'][idx]):.6f}")
    lines.append("")
    lines.append("--- 边界构型指定把手位置与速度 ---")
    lines.append("部位\tx(m)\ty(m)\t速度(m/s)")
    for idx in selected:
        p = boundary["positions"][idx]
        lines.append(f"{handle_label(idx)}\t{p[0]:.6f}\t{p[1]:.6f}"
                     f"\t{abs(boundary['speeds'][idx]):.6f}")

    txt_path = os.path.join(BASE, "表4_最小螺距关键结果.txt")
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    print("saved", txt_path)


def draw_margin_curve(result: dict) -> None:
    pitch = result["pitch"]
    grid = np.linspace(0.32, 0.90, 41)
    margins = np.empty(grid.shape)
    pairs = []
    for k, p in enumerate(grid):
        state = path_clearance_for_pitch(float(p), grid_size=21)
        margins[k] = state["clearance"]
        pairs.append(state["pair"])
    print("margin curve p range:", float(grid[0]), float(grid[-1]))
    print("min margin on curve grid:", float(margins.min()), "at p =",
          float(grid[int(np.argmin(margins))]))

    fig, ax = plt.subplots(figsize=(9.0, 5.6))
    ax.axhline(0.0, color="#7f8c8d", linewidth=1.0, linestyle="--")
    ax.plot(grid, margins, "-", color="#2c7fb8", linewidth=2.0,
            label="全过程最小安全裕量 H(p)")
    ax.axvline(pitch, color="#c0392b", linewidth=1.4, linestyle="--",
               label=f"p_min = {pitch:.6f} m")
    ax.fill_between(grid, 0.0, margins, where=margins >= 0,
                    color="#27ae60", alpha=0.12, label="可行区 H(p)≥0")
    ax.fill_between(grid, 0.0, margins, where=margins < 0,
                    color="#e74c3c", alpha=0.12, label="不可行区 H(p)<0")
    ax.plot([pitch], [0.0], "o", color="#c0392b", markersize=8)
    ax.set_xlabel("螺距 p (m)")
    ax.set_ylabel("全过程最小安全裕量 H(p) (m)")
    ax.set_title("第三问 安全裕量随螺距变化与最小螺距定位")
    ax.grid(True, linestyle=":", linewidth=0.6, alpha=0.6)
    ax.legend(loc="upper left", fontsize=9)
    fig.tight_layout()
    png = os.path.join(BASE, "螺距_安全裕量曲线.png")
    fig.savefig(png, dpi=170)
    plt.close(fig)
    print("saved", png)


def draw_critical_configuration(result: dict) -> None:
    pitch = result["pitch"]
    path_min = result["path_min"]
    critical = result["critical"]
    pair = path_min["pair"]
    points = critical["positions"]
    theta_head = path_min["theta"]
    r_head = theta_head * spiral_b(pitch)

    fig, ax = plt.subplots(figsize=(9.5, 9.5))
    for k in range(1, N_BOARDS + 1):
        corners = board_rectangle(points, k)
        if k in pair:
            color = "#e74c3c"
            alpha = 0.80
            zorder = 5
        else:
            color = "#2c7fb8"
            alpha = 0.14
            zorder = 1
        ax.add_patch(Polygon(corners, closed=True, fill=True, facecolor=color,
                             edgecolor="none", alpha=alpha, zorder=zorder))

    theta_tail = theta_head + 8.0
    ref = np.linspace(theta_head - 0.2, theta_tail, 600)
    ax.plot([spiral_point(th, pitch)[0] for th in ref],
            [spiral_point(th, pitch)[1] for th in ref],
            "-", color="#34495e", linewidth=0.7, alpha=0.7, label="等距螺线参考线")

    ax.plot(points[:, 0], points[:, 1], "-", color="#1f4e79", linewidth=0.9,
            label="把手中心连线", zorder=3)
    ax.plot(points[0, 0], points[0, 1], "o", color="#c0392b", markersize=7,
            label="龙头前把手", zorder=6)
    ax.add_patch(Circle((0, 0), TURN_RADIUS, fill=False, color="#8e44ad",
                        linestyle="--", linewidth=1.6, label="调头空间边界 R=4.5 m"))
    ax.plot(points[pair[0] - 1, 0], points[pair[0] - 1, 1], "s", color="#e67e22",
            markersize=9, label=f"临界板 {pair[0]} 前端", zorder=6)
    ax.plot(points[pair[0], 0], points[pair[0], 1], "s", color="#e67e22",
            markersize=9, zorder=6)
    ax.plot(points[pair[1] - 1, 0], points[pair[1] - 1, 1], "^", color="#8e44ad",
            markersize=9, label=f"临界板 {pair[1]} 前端", zorder=6)
    ax.plot(points[pair[1], 0], points[pair[1], 1], "^", color="#8e44ad",
            markersize=9, zorder=6)

    ax.set_aspect("equal", adjustable="box")
    ax.grid(True, linestyle=":", linewidth=0.6, alpha=0.6)
    ax.set_title(f"第三问临界构型（p_min={pitch:.6f} m, 龙头极径 r={r_head:.4f} m）")
    ax.set_xlabel("x (m)")
    ax.set_ylabel("y (m)")
    ax.legend(loc="best", fontsize=9)
    fig.tight_layout()
    png = os.path.join(BASE, "临界构型_碰撞示意图.png")
    fig.savefig(png, dpi=170)
    plt.close(fig)
    print("saved", png)


def main() -> None:
    low = 4.5 / 16 + 1e-6
    high = 0.80
    search = bisect_minimum_pitch(low, high)
    pitch = search["pitch"]

    boundary_clearance, boundary_pair, theta_boundary, boundary_points = (
        boundary_clearance_for_pitch(pitch)
    )
    path_min = path_clearance_for_pitch(pitch, grid_size=81)

    _, boundary_positions, boundary_tangents, boundary_speeds = spiral_state(
        theta_boundary, pitch, 1.0
    )
    _, critical_positions, critical_tangents, critical_speeds = spiral_state(
        path_min["theta"], pitch, 1.0
    )

    below = path_clearance_for_pitch(pitch - 1e-3)["clearance"]
    above = path_clearance_for_pitch(pitch + 1e-3)["clearance"]
    residual = constraint_residual(boundary_positions)

    result = {
        "pitch": pitch,
        "history": search["history"],
        "boundary": {
            "theta": theta_boundary,
            "clearance": boundary_clearance,
            "pair": boundary_pair,
            "positions": boundary_positions,
            "tangents": boundary_tangents,
            "speeds": boundary_speeds,
        },
        "path_min": {
            "clearance": path_min["clearance"],
            "theta": path_min["theta"],
            "pair": path_min["pair"],
        },
        "critical": {
            "positions": critical_positions,
            "tangents": critical_tangents,
            "speeds": critical_speeds,
        },
        "verification": {"below": below, "above": above},
        "residual": residual,
    }

    write_result3(result)
    write_key_txt(result)
    draw_margin_curve(result)
    draw_critical_configuration(result)

    print("p_min =", f"{pitch:.9f}")
    print("H(p_min) =", f"{path_min['clearance']:.6e}")
    print("critical pair =", path_min["pair"], "theta =", f"{path_min['theta']:.9f}")
    print("boundary clearance =", f"{boundary_clearance:.6e}",
          "pair =", boundary_pair)
    print("verification below/above =",
          f"{below:.6e}", f"{above:.6e}")


if __name__ == "__main__":
    main()
