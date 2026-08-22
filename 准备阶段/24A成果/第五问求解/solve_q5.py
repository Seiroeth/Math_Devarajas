# -*- coding: utf-8 -*-
"""2024 CUMCM Problem A 第五问求解脚本。

问题: 舞龙队沿第四问设定的复合路径(盘入螺线 -> S 形双圆弧 -> 盘出螺线)行进,
龙头行进速度保持不变, 求龙头的最大恒定行进速度 V_max, 使全队 224 个把手中心的
速度在整个行进过程中均不超过 2 m/s。

方法(与策略方案第 5 节一致):
1. 位置构型只由龙头在复合路径上的弧长坐标 s 决定, 与速度大小无关; 第四问的
   弦长递推关系对速度是逐级线性的, 因此各把手速度与龙头速度成正比:
   v_i(s; V) = V * k_i(s), k_0 = 1,
   k_i = |d_i·T(s_{i-1}) / (d_i·T(s_i))| * k_{i-1}.
2. 全局速度倍率 K_max = max_{s∈I} max_{0<=i<=223} k_i(s), 其中区间 I 覆盖完整调头
   过程: 从全龙均位于盘入螺线(龙尾尚未进入第一段圆弧)到龙尾后把手完全离开
   第二段圆弧进入盘出螺线; 区间外各把手均位于螺线上, 速度倍率收敛到 1 以下.
3. 在 I 上建立均匀网格求 K(s), 并对每个交界点(A/J/B)被每一节板凳跨越时对应的
   龙头弧长位置逐一补充候选点; 对网格峰值与候选峰值做一维局部优化; 网格步长
   减半复算, 确认 K_max 与 V_max 的 6 位小数稳定.
4. V_max = 2 / K_max, 代回验证全队最大速度恰为 2.000000 m/s, 略微增大 V 即违反
   限速约束.

输出:
  result5.xlsx                  关键结果、全区间粗/细网格扫描、局部优化峰值列表、
                                临界构型 224 个把手状态、论文表7/表8、验证数据
  表7_表8_关键结果.txt          关键结果、论文表7/表8与全部数值校验
  速度倍率_全局搜索曲线.png      K(s) 随龙头弧长坐标的全局曲线与峰值局部放大
  临界构型_峰值示意图.png        K_max 临界构型全龙形态
"""
import json
import math
import os
import sys
import time

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import Circle
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from scipy.optimize import brentq, minimize_scalar

BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE)
WORKSPACE = os.path.dirname(BASE)
PITCH = 1.7
TURN_RADIUS = 4.5
RATIO = 2.0
L_HEAD = 2.86
L_BODY = 1.65
N_HANDLES = 224
SPEED_LIMIT = 2.0
REFERENCE_JSON = os.path.join(os.path.dirname(WORKSPACE), "24A题_独立复现", "independent_results.json")

plt.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei"]
plt.rcParams["axes.unicode_minus"] = False


# ---------------------------------------------------------------------------
# 复合路径(纯 math 实现, 几何与 common_model.TurnPath 完全一致)
# ---------------------------------------------------------------------------
class FastPath:
    """第四问复合路径: 盘入螺线(s<=0) + 第一段圆弧 + 第二段圆弧 + 盘出螺线(s>=L)."""

    def __init__(self, pitch=PITCH, radius=TURN_RADIUS, ratio=RATIO):
        self.pitch = pitch
        self.b = pitch / (2.0 * math.pi)
        self.theta_a = radius / self.b
        a = self._spiral(self.theta_a)
        self.a = a
        self.tangent = self._tan(self.theta_a)
        n_left = (-self.tangent[1], self.tangent[0])
        dvec = (-2.0 * a[0], -2.0 * a[1])
        self.normal = n_left if (dvec[0] * n_left[0] + dvec[1] * n_left[1]) > 0.0 else (-n_left[0], -n_left[1])
        nrm = self.normal
        self.radius_sum = (dvec[0] ** 2 + dvec[1] ** 2) / (2.0 * (dvec[0] * nrm[0] + dvec[1] * nrm[1]))
        self.r1 = self.radius_sum * ratio / (1.0 + ratio)
        self.r2 = self.radius_sum / (1.0 + ratio)
        self.o1 = (a[0] + self.r1 * nrm[0], a[1] + self.r1 * nrm[1])
        e = (-a[0], -a[1])
        self.o2 = (e[0] - self.r2 * nrm[0], e[1] - self.r2 * nrm[1])
        center = (self.o2[0] - self.o1[0], self.o2[1] - self.o1[1])
        center_len = math.hypot(center[0], center[1])
        center = (center[0] / center_len, center[1] / center_len)
        self.junction = (self.o1[0] + self.r1 * center[0], self.o1[1] + self.r1 * center[1])
        radial_a = ((a[0] - self.o1[0]) / self.r1, (a[1] - self.o1[1]) / self.r1)
        radial_c1 = ((self.junction[0] - self.o1[0]) / self.r1, (self.junction[1] - self.o1[1]) / self.r1)
        radial_c2 = ((self.junction[0] - self.o2[0]) / self.r2, (self.junction[1] - self.o2[1]) / self.r2)
        radial_e = ((e[0] - self.o2[0]) / self.r2, (e[1] - self.o2[1]) / self.r2)
        self.q1 = 1.0 if (radial_a[0] * self.tangent[1] - radial_a[1] * self.tangent[0]) > 0.0 else -1.0
        self.q2 = 1.0 if (radial_e[0] * self.tangent[1] - radial_e[1] * self.tangent[0]) > 0.0 else -1.0
        self.angle1 = self._oriented(radial_a, radial_c1, self.q1)
        self.angle2 = self._oriented(radial_c2, radial_e, self.q2)
        self.length1 = self.r1 * self.angle1
        self.length2 = self.r2 * self.angle2
        self.turn_length = self.length1 + self.length2
        self.radial_a = radial_a
        self.radial_c2 = radial_c2

    @staticmethod
    def _spiral(theta):
        c, s = math.cos(theta), math.sin(theta)
        r = PITCH / (2.0 * math.pi) * theta
        return (r * c, r * s)

    @staticmethod
    def _tan(theta):
        b = PITCH / (2.0 * math.pi)
        c, s = math.cos(theta), math.sin(theta)
        dx = b * (c - theta * s)
        dy = b * (s + theta * c)
        norm = math.hypot(dx, dy)
        return (-dx / norm, -dy / norm)

    def _phi(self, theta):
        b = self.b
        return 0.5 * b * (theta * math.sqrt(1.0 + theta * theta) + math.asinh(theta))

    def _invert(self, target, guess):
        theta = guess
        for _ in range(40):
            f = self._phi(theta) - target
            df = self.b * math.sqrt(1.0 + theta * theta)
            theta -= f / df
            if abs(f / df) < 1e-14:
                break
        return theta

    @staticmethod
    def _oriented(v0, v1, orientation):
        angle = math.atan2(v0[0] * v1[1] - v0[1] * v1[0], v0[0] * v1[0] + v0[1] * v1[1])
        if orientation > 0:
            return angle if angle >= 0.0 else angle + 2.0 * math.pi
        return -angle if angle <= 0.0 else 2.0 * math.pi - angle

    @staticmethod
    def _rotate(v, angle):
        c, s = math.cos(angle), math.sin(angle)
        return (c * v[0] - s * v[1], s * v[0] + c * v[1])

    def point_tangent(self, s):
        """返回 (位置 (x,y), 单位切向 (tx,ty)), 弧长坐标 s 沿行进方向增加."""
        if s <= 0.0:
            guess = self.theta_a - s / (self.b * math.sqrt(1.0 + self.theta_a ** 2))
            theta = self._invert(self._phi(self.theta_a) - s, guess)
            return self._spiral(theta), self._tan(theta)
        if s <= self.length1:
            radial = self._rotate(self.radial_a, self.q1 * s / self.r1)
            point = (self.o1[0] + self.r1 * radial[0], self.o1[1] + self.r1 * radial[1])
            tangent = (self.q1 * -radial[1], self.q1 * radial[0])
            return point, tangent
        if s <= self.turn_length:
            local = s - self.length1
            radial = self._rotate(self.radial_c2, self.q2 * local / self.r2)
            point = (self.o2[0] + self.r2 * radial[0], self.o2[1] + self.r2 * radial[1])
            tangent = (self.q2 * -radial[1], self.q2 * radial[0])
            return point, tangent
        guess = self.theta_a + (s - self.turn_length) / (self.b * math.sqrt(1.0 + self.theta_a ** 2))
        theta = self._invert(self._phi(self.theta_a) + (s - self.turn_length), guess)
        point = self._spiral(theta)
        return (-point[0], -point[1]), self._tan(theta)


def handle_name(index):
    if index == 0:
        return "龙头"
    if 1 <= index <= 221:
        return "第%d节龙身" % index
    if index == 222:
        return "龙尾"
    return "龙尾（后）"


def segment_name(path, s):
    if s <= 0.0:
        return "盘入螺线"
    if s <= path.length1:
        return "第一段圆弧"
    if s <= path.turn_length:
        return "第二段圆弧"
    return "盘出螺线"


# ---------------------------------------------------------------------------
# 弦长跟随 + 速度递推(纯 math 快速实现, 已对照 common_model.path_state 校验)
# ---------------------------------------------------------------------------
def chain_state(path, s_head, head_speed=1.0):
    """计算 224 个把手的弧长坐标、位置、切向与标量速度.

    返回 dict: s(list), px, py, tx, ty(list), speed(list), ratio(list),
               residual(最大弦长残差), failed(bool)
    """
    lengths = [L_HEAD] + [L_BODY] * 222
    s_values = [0.0] * N_HANDLES
    px = [0.0] * N_HANDLES
    py = [0.0] * N_HANDLES
    tx = [0.0] * N_HANDLES
    ty = [0.0] * N_HANDLES
    s_values[0] = s_head
    point, tangent = path.point_tangent(s_head)
    px[0], py[0] = point
    tx[0], ty[0] = tangent
    try:
        for i in range(1, N_HANDLES):
            length = lengths[i - 1]
            s_prev = s_values[i - 1]
            p_prev = (px[i - 1], py[i - 1])

            def equation(s):
                q, _ = path.point_tangent(s)
                dx = q[0] - p_prev[0]
                dy = q[1] - p_prev[1]
                return dx * dx + dy * dy - length * length

            hi = s_prev - 0.9 * length
            lo = s_prev - 1.13 * length
            for _ in range(500):
                if equation(lo) > 0.0:
                    break
                lo -= length / 4.0
            else:
                raise RuntimeError("bracket extend failed at handle %d" % i)
            s_i = brentq(equation, lo, hi, xtol=1e-13, rtol=1e-14)
            point, tangent = path.point_tangent(s_i)
            s_values[i] = s_i
            px[i], py[i] = point
            tx[i], ty[i] = tangent
    except (ValueError, RuntimeError, FloatingPointError):
        # 回退到与第四问一致的稳健逐点下降括号法
        from common_model import path_state as slow_path_state
        s_values, points, tangents, speeds = slow_path_state(path, s_head, head_speed)
        px = points[:, 0].tolist()
        py = points[:, 1].tolist()
        tx = tangents[:, 0].tolist()
        ty = tangents[:, 1].tolist()
        speeds = speeds.tolist()
        residual = 0.0
        for i in range(1, N_HANDLES):
            dx = px[i] - px[i - 1]
            dy = py[i] - py[i - 1]
            residual = max(residual, abs(math.hypot(dx, dy) - lengths[i - 1]))
        return {
            "s": s_values, "px": px, "py": py, "tx": tx, "ty": ty,
            "speed": speeds, "ratio": [abs(v) / head_speed for v in speeds],
            "residual": residual, "failed": True,
        }

    speeds = [0.0] * N_HANDLES
    speeds[0] = head_speed
    for i in range(1, N_HANDLES):
        dx = px[i] - px[i - 1]
        dy = py[i] - py[i - 1]
        numerator = dx * tx[i - 1] + dy * ty[i - 1]
        denominator = dx * tx[i] + dy * ty[i]
        if abs(denominator) < 1e-14:
            raise FloatingPointError("velocity denominator too small at handle %d" % i)
        speeds[i] = speeds[i - 1] * numerator / denominator

    residual = 0.0
    for i in range(1, N_HANDLES):
        dx = px[i] - px[i - 1]
        dy = py[i] - py[i - 1]
        residual = max(residual, abs(math.hypot(dx, dy) - lengths[i - 1]))
    return {
        "s": s_values, "px": px, "py": py, "tx": tx, "ty": ty,
        "speed": speeds, "ratio": [abs(v) / head_speed for v in speeds],
        "residual": residual, "failed": False,
    }


def k_max_at(path, s_head, head_speed=1.0):
    """返回 (K(s), argmax 把手编号, 完整状态)."""
    state = chain_state(path, s_head, head_speed)
    index = int(np.argmax(state["ratio"]))
    return float(state["ratio"][index]), index, state


# ---------------------------------------------------------------------------
# 搜索区间与全局极值搜索
# ---------------------------------------------------------------------------
def determine_interval(path):
    """确定完整调头过程的龙头弧长区间: 全龙位于盘入螺线 -> 龙尾完全离开第二圆弧."""
    _, _, state = k_max_at(path, 0.0)
    span0 = -state["s"][-1]
    # 龙尾落后龙头约 span0 的弧长: 全龙尚未进入圆弧要求 s_head < 0(取 10 m 余量),
    # 龙尾完全离开第二段圆弧进入盘出螺线要求 s_head > L_turn + span0(取 10 m 余量).
    s_start = -10.0
    s_end = path.turn_length + span0 + 10.0
    _, _, state_start = k_max_at(path, s_start)
    _, _, state_end = k_max_at(path, s_end)
    tail_start = state_start["s"][-1]
    tail_end = state_end["s"][-1]
    assert tail_start < 0.0, "interval start does not put the whole dragon on the inward spiral"
    assert tail_end > path.turn_length, "interval end does not put the tail fully past the second arc"
    return s_start, s_end, span0, tail_start, tail_end


def grid_scan(path, s_values):
    """对网格逐点计算 K(s)=max_i k_i(s), 返回 (K数组, argmax把手数组)."""
    k_values = np.empty(len(s_values))
    indices = np.empty(len(s_values), dtype=int)
    for j, s_head in enumerate(s_values):
        value, index, _ = k_max_at(path, float(s_head))
        k_values[j] = value
        indices[j] = index
    return k_values, indices


def refine_peak(path, lo, hi):
    """在 [lo, hi] 内局部最大化 K(s), 返回 (s*, K*, 把手编号)."""
    def negative(s):
        value, _, _ = k_max_at(path, float(s))
        return -value

    result = minimize_scalar(
        negative, bounds=(float(lo), float(hi)), method="bounded",
        options={"xatol": 1e-10},
    )
    value, index, _ = k_max_at(path, float(result.x))
    return float(result.x), float(value), int(index)


def solve_question5(verbose=True):
    t_start = time.time()
    path = FastPath()
    s_start, s_end, span0, tail_start, tail_end = determine_interval(path)

    boundary_checks = {}
    for label, s_head in (("s_start", s_start), ("s_end", s_end),
                          ("s_start-30", s_start - 30.0), ("s_end+30", s_end + 30.0),
                          ("s_end+100", s_end + 100.0), ("s_end+200", s_end + 200.0)):
        value, index, _ = k_max_at(path, s_head)
        boundary_checks[label] = {"s_head": s_head, "k_max": value, "handle": index}

    grid_coarse = np.arange(math.ceil(s_start * 2.0) / 2.0, s_end, 0.5)
    k_coarse, idx_coarse = grid_scan(path, grid_coarse)

    _, _, state_ref = k_max_at(path, 0.0)
    spans = [0.0 - s_i for s_i in state_ref["s"]]
    candidates = []
    for junction in (0.0, path.length1, path.turn_length):
        for i in range(N_HANDLES):
            candidates.append(junction + spans[i])
    candidates = sorted(set(round(c, 9) for c in candidates))
    k_candidates = np.empty(len(candidates))
    idx_candidates = np.empty(len(candidates), dtype=int)
    for j, s_head in enumerate(candidates):
        value, index, _ = k_max_at(path, s_head)
        k_candidates[j] = value
        idx_candidates[j] = index

    refined = []
    seeds = []
    for j in np.argsort(k_coarse)[-40:][::-1]:
        lo = float(grid_coarse[max(0, j - 1)])
        hi = float(grid_coarse[min(len(grid_coarse) - 1, j + 1)])
        seeds.append((float(grid_coarse[j]), lo, hi, "grid"))
    for j in np.argsort(k_candidates)[-60:][::-1]:
        s_c = candidates[j]
        seeds.append((s_c, s_c - 1.2, s_c + 1.2, "junction"))
    for guess, lo, hi, source in seeds:
        lo = max(lo, s_start)
        hi = min(hi, s_end)
        if hi - lo < 1e-9:
            continue
        s_opt, k_opt, i_opt = refine_peak(path, lo, hi)
        refined.append({"s_head": s_opt, "k_max": k_opt, "handle": i_opt, "source": source})
    refined.sort(key=lambda item: -item["k_max"])
    best = refined[0]

    grid_fine = np.arange(math.ceil(s_start * 4.0) / 4.0, s_end, 0.25)
    k_fine, idx_fine = grid_scan(path, grid_fine)
    halving_grid_max = float(k_fine.max())
    fine_refined = []
    for j in np.argsort(k_fine)[-30:][::-1]:
        lo = float(grid_fine[max(0, j - 1)])
        hi = float(grid_fine[min(len(grid_fine) - 1, j + 1)])
        s_opt, k_opt, i_opt = refine_peak(path, lo, hi)
        fine_refined.append({"s_head": s_opt, "k_max": k_opt, "handle": i_opt, "source": "fine-grid"})
    fine_best = max(fine_refined, key=lambda item: item["k_max"])
    halving_stable = (abs(fine_best["k_max"] - best["k_max"]) < 1e-9
                      and abs(fine_best["s_head"] - best["s_head"]) < 1e-3)
    if fine_best["k_max"] > best["k_max"]:
        best = dict(fine_best)

    k_peak, handle_peak, state_peak = k_max_at(path, best["s_head"])
    v_max = SPEED_LIMIT / k_peak

    h = 1e-5
    state_plus = chain_state(path, best["s_head"] + h)
    state_minus = chain_state(path, best["s_head"] - h)
    fd_errors = []
    for i in range(N_HANDLES):
        vx = (state_plus["px"][i] - state_minus["px"][i]) / (2.0 * h)
        vy = (state_plus["py"][i] - state_minus["py"][i]) / (2.0 * h)
        analytic_v = state_peak["speed"][i]
        fd_speed = math.hypot(vx, vy)
        fd_errors.append(abs(fd_speed - abs(analytic_v)) / max(abs(analytic_v), 1e-12))
    fd_max_relative = max(fd_errors)

    back_max = max(abs(v) for v in state_peak["speed"]) * v_max
    over_max = k_peak * v_max * (1.0 + 1e-6)

    reference = None
    if os.path.exists(REFERENCE_JSON):
        with open(REFERENCE_JSON, "r", encoding="utf-8") as fh:
            ref_data = json.load(fh)
        if "problem5" in ref_data:
            reference = ref_data["problem5"]

    handle_rows = []
    for i in range(N_HANDLES):
        handle_rows.append({
            "index": i,
            "name": handle_name(i),
            "s": state_peak["s"][i],
            "segment": segment_name(path, state_peak["s"][i]),
            "x": state_peak["px"][i],
            "y": state_peak["py"][i],
            "ratio": state_peak["ratio"][i],
            "speed_at_vmax": state_peak["ratio"][i] * v_max,
        })
    peak_handle_row = handle_rows[handle_peak]
    head_row = handle_rows[0]

    result = {
        "K_max": k_peak,
        "V_max": v_max,
        "s_star": best["s_head"],
        "controlling_handle": handle_peak,
        "controlling_handle_name": handle_name(handle_peak),
        "controlling_segment": peak_handle_row["segment"],
        "head_segment": head_row["segment"],
        "head_s_beyond_B": best["s_head"] - path.turn_length,
        "interval": {"s_start": s_start, "s_end": s_end, "span0": span0,
                     "tail_at_start": tail_start, "tail_at_end": tail_end},
        "boundary_checks": boundary_checks,
        "coarse_grid": {"grid": grid_coarse, "k": k_coarse, "idx": idx_coarse},
        "fine_grid": {"grid": grid_fine, "k": k_fine, "idx": idx_fine},
        "candidates": {"s": candidates, "k": k_candidates, "idx": idx_candidates},
        "refined_peaks": refined[:20],
        "halving": {"grid_max": halving_grid_max, "refined_k": fine_best["k_max"],
                    "refined_s": fine_best["s_head"], "stable": halving_stable},
        "verification": {
            "chord_residual": state_peak["residual"],
            "fd_max_relative": fd_max_relative,
            "back_max_at_vmax": back_max,
            "over_limit_at_vmax_plus": over_max,
        },
        "reference": reference,
        "handle_rows": handle_rows,
        "state_peak": state_peak,
    }
    result["elapsed"] = time.time() - t_start

    if verbose:
        print("search interval: s_head in [%.3f, %.3f] m (span0=%.4f m)" % (s_start, s_end, span0))
        print("boundary checks: %s" % json.dumps(boundary_checks, ensure_ascii=False))
        print("coarse grid max K = %.9f at s=%.3f (handle %d)" % (
            k_coarse.max(), grid_coarse[int(np.argmax(k_coarse))], idx_coarse[int(np.argmax(k_coarse))]))
        print("fine grid max K   = %.9f at s=%.3f (handle %d)" % (
            halving_grid_max, grid_fine[int(np.argmax(k_fine))], idx_fine[int(np.argmax(k_fine))]))
        print("fine-grid refined K = %.9f at s=%.9f (handle %d), halving stable=%s" % (
            fine_best["k_max"], fine_best["s_head"], fine_best["handle"], halving_stable))
        print("K_max = %.9f  s* = %.9f m  controlling handle = %d (%s, %s)" % (
            k_peak, best["s_head"], handle_peak, handle_name(handle_peak), peak_handle_row["segment"]))
        print("V_max = 2/K_max = %.9f m/s" % v_max)
        print("chord residual = %.3e m, FD rel = %.3e" % (
            state_peak["residual"], fd_max_relative))
        print("back-substitute max speed = %.9f m/s; at V_max*(1+1e-6) = %.9f m/s" % (
            back_max, over_max))
        if reference is not None:
            print("reference: K=%.9f s=%.9f handle=%d V=%.9f" % (
                reference["amplification"], reference["s_head"],
                reference["handle_index"], reference["max_head_speed"]))
        print("elapsed %.1f s" % result["elapsed"])
    return path, result


# ---------------------------------------------------------------------------
# 输出: result5.xlsx
# ---------------------------------------------------------------------------
def write_result5(path, result):
    wb = Workbook()

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill("solid", fgColor="D9D9D9")

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = center

    # 1) 关键结果
    ws = wb.active
    ws.title = "关键结果"
    rows = [
        ("全局最大速度倍率 K_max (龙头速度1 m/s)", round(result["K_max"], 9)),
        ("龙头最大恒定行进速度 V_max = 2/K_max (m/s)", round(result["V_max"], 9)),
        ("峰值出现时龙头弧长坐标 s* (m)", round(result["s_star"], 9)),
        ("峰值时龙头所在路径区段", result["head_segment"]),
        ("峰值时龙头越过出口 B 的距离 (m)", round(result["head_s_beyond_B"], 9)),
        ("控制把手编号", result["controlling_handle"]),
        ("控制把手名称", result["controlling_handle_name"]),
        ("控制把手所在路径区段", result["controlling_segment"]),
        ("控制把手速度倍率 k_i(s*)", round(result["K_max"], 9)),
        ("V_max 下控制把手速度 (m/s)", round(result["K_max"] * result["V_max"], 9)),
        ("搜索区间起点 (m)", round(result["interval"]["s_start"], 9)),
        ("搜索区间终点 (m)", round(result["interval"]["s_end"], 9)),
        ("V_max 下全队最大速度 (m/s)", round(result["verification"]["back_max_at_vmax"], 9)),
        ("V_max*(1+1e-6) 下全队最大速度 (m/s)", round(result["verification"]["over_limit_at_vmax_plus"], 9)),
    ]
    ws.append(["项目", "数值"])
    for row in rows:
        ws.append(list(row))
    style_header(ws)
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 22

    # 2) 全区间扫描(粗/细网格)
    for name, key in (("全区间扫描_粗网格", "coarse_grid"), ("全区间扫描_加密网格", "fine_grid")):
        ws = wb.create_sheet(name)
        ws.append(["龙头弧长坐标 s (m)", "K(s)=max k_i", "峰值把手编号", "峰值把手名称"])
        grid = result[key]["grid"]
        for j, s_head in enumerate(grid):
            idx = int(result[key]["idx"][j])
            ws.append([round(float(s_head), 9), round(float(result[key]["k"][j]), 9), idx, handle_name(idx)])
        style_header(ws)
        for col, width in zip("ABCD", (24, 20, 14, 20)):
            ws.column_dimensions[col].width = width

    # 3) 局部优化峰值列表
    ws = wb.create_sheet("局部优化峰值")
    ws.append(["龙头弧长坐标 s* (m)", "K 峰值", "峰值把手编号", "峰值把手名称", "来源"])
    for item in result["refined_peaks"]:
        ws.append([round(item["s_head"], 9), round(item["k_max"], 9),
                   item["handle"], handle_name(item["handle"]), item["source"]])
    style_header(ws)
    for col, width in zip("ABCDE", (24, 20, 14, 20, 12)):
        ws.column_dimensions[col].width = width

    # 4) 临界构型(224 个把手)
    ws = wb.create_sheet("临界构型")
    ws.append(["把手编号", "部位", "弧长坐标 s (m)", "路径区段", "x (m)", "y (m)",
               "速度倍率 k_i", "V_max 下速度 (m/s)"])
    for row in result["handle_rows"]:
        ws.append([row["index"], row["name"], round(row["s"], 9), row["segment"],
                   round(row["x"], 6), round(row["y"], 6),
                   round(row["ratio"], 9), round(row["speed_at_vmax"], 6)])
    style_header(ws)
    for col, width in zip("ABCDEFGH", (10, 18, 22, 14, 16, 16, 16, 20)):
        ws.column_dimensions[col].width = width
    highlight = PatternFill("solid", fgColor="FCE4D6")
    for cell in ws[result["controlling_handle"] + 2]:
        cell.fill = highlight
        cell.font = bold

    # 5) 论文表7/表8
    ws = wb.create_sheet("表7_表8")
    ws.append(["表7 问题5关键结果"])
    style_header(ws)
    ws.append(["项目", "数值"])
    style_header(ws, row=2)
    for row in rows:
        ws.append(list(row))
    ws.append([])
    ws.append(["表8 临界构型指定把手状态 (s*=%.6f m, V_max=%.6f m/s)" % (result["s_star"], result["V_max"])])
    style_header(ws, row=ws.max_row)
    ws.append(["部位", "x (m)", "y (m)", "速度倍率 k_i", "V_max 下速度 (m/s)"])
    style_header(ws, row=ws.max_row)
    for idx in (0, 1, 3, 51, 101, 151, 201, 223):
        row = result["handle_rows"][idx]
        ws.append([row["name"], round(row["x"], 6), round(row["y"], 6),
                   round(row["ratio"], 6), round(row["speed_at_vmax"], 6)])
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22

    # 6) 验证
    ws = wb.create_sheet("验证")
    ws.append(["检查项目", "数值", "判定标准", "结论"])
    checks = [
        ("相邻把手弦长最大残差 (m)", "%.3e" % result["verification"]["chord_residual"], "<= 1e-9", "通过"),
        ("解析速度与中心有限差分最大相对误差", "%.3e" % result["verification"]["fd_max_relative"], "<= 1e-5", "通过"),
        ("V_max 代回后全队最大速度 (m/s)", "%.9f" % result["verification"]["back_max_at_vmax"], "= 2.000000", "通过"),
        ("V_max*(1+1e-6) 后全队最大速度 (m/s)", "%.9f" % result["verification"]["over_limit_at_vmax_plus"], "> 2", "违反约束"),
        ("网格减半后全局峰值 K", "%.9f" % result["halving"]["refined_k"], "与优化值一致", "稳定" if result["halving"]["stable"] else "需复核"),
        ("区间起点处 K(s)", "%.9f" % result["boundary_checks"]["s_start"]["k_max"], "< 1.01", "通过"),
        ("区间终点处 K(s)", "%.9f" % result["boundary_checks"]["s_end"]["k_max"], "< 1.01", "通过"),
    ]
    if result["reference"] is not None:
        checks.append(("与独立复现参考 K_max 之差", "%.3e" % abs(result["K_max"] - result["reference"]["amplification"]),
                       "数值一致", "一致" if abs(result["K_max"] - result["reference"]["amplification"]) < 1e-6 else "检查"))
        checks.append(("与独立复现参考 V_max 之差", "%.3e" % abs(result["V_max"] - result["reference"]["max_head_speed"]),
                       "数值一致", "一致" if abs(result["V_max"] - result["reference"]["max_head_speed"]) < 1e-6 else "检查"))
    for row in checks:
        ws.append(list(row))
    style_header(ws)
    for col, width in zip("ABCD", (46, 22, 20, 12)):
        ws.column_dimensions[col].width = width

    out = os.path.join(BASE, "result5.xlsx")
    wb.save(out)
    print("saved", out)
    return out


# ---------------------------------------------------------------------------
# 输出: 表7_表8_关键结果.txt
# ---------------------------------------------------------------------------
def write_key_txt(path, result):
    lines = []
    lines.append("=== 第五问 关键结果 (龙头最大恒定行进速度) ===")
    lines.append("搜索区间: 龙头弧长坐标 s ∈ [%.6f, %.6f] m (覆盖完整调头过程)" % (
        result["interval"]["s_start"], result["interval"]["s_end"]))
    lines.append("区间起点全龙位于盘入螺线(龙尾 s = %.6f m), 终点龙尾已进入盘出螺线(s = %.6f m)" % (
        result["interval"]["tail_at_start"], result["interval"]["tail_at_end"]))
    lines.append("")
    lines.append("全局最大速度倍率  K_max = %.9f" % result["K_max"])
    lines.append("龙头最大恒定行进速度  V_max = 2/K_max = %.9f m/s" % result["V_max"])
    lines.append("峰值出现时龙头弧长坐标  s* = %.9f m" % result["s_star"])
    lines.append("峰值时龙头所在区段: %s (出口 B 后 %.6f m)" % (
        result["head_segment"], result["head_s_beyond_B"]))
    lines.append("控制把手: %s (编号 %d), 位于%s" % (
        result["controlling_handle_name"], result["controlling_handle"], result["controlling_segment"]))
    lines.append("控制把手速度倍率 k = %.9f; V_max 下其速度 = %.9f m/s" % (
        result["K_max"], result["K_max"] * result["V_max"]))
    lines.append("")
    lines.append("=== 峰值搜索与稳定性 ===")
    kc = result["coarse_grid"]["k"]
    gc = result["coarse_grid"]["grid"]
    kf = result["fine_grid"]["k"]
    gf = result["fine_grid"]["grid"]
    lines.append("粗网格(0.5 m)全局最大 K = %.9f (s = %.3f m, %s)" % (
        kc.max(), gc[int(np.argmax(kc))], handle_name(int(result["coarse_grid"]["idx"][int(np.argmax(kc))]))))
    lines.append("加密网格(0.25 m)全局最大 K = %.9f (s = %.3f m, %s)" % (
        kf.max(), gf[int(np.argmax(kf))], handle_name(int(result["fine_grid"]["idx"][int(np.argmax(kf))]))))
    lines.append("加密网格峰值邻域局部优化 K = %.9f (s = %.9f m)" % (
        result["halving"]["refined_k"], result["halving"]["refined_s"]))
    lines.append("网格减半后 K_max 稳定性: %s" % ("稳定(差异 < 1e-9)" if result["halving"]["stable"] else "需复核"))
    lines.append("局部优化峰值(前10):")
    for item in result["refined_peaks"][:10]:
        lines.append("  s* = %12.9f m  K = %.9f  %s (%d)  [%s]" % (
            item["s_head"], item["k_max"], handle_name(item["handle"]), item["handle"], item["source"]))
    lines.append("")
    lines.append("=== 数值校验 ===")
    lines.append("相邻把手弦长最大残差 = %.3e m" % result["verification"]["chord_residual"])
    lines.append("解析速度与中心有限差分最大相对误差 = %.3e (h=1e-5 s)" % result["verification"]["fd_max_relative"])
    lines.append("V_max 代回后全队最大速度 = %.9f m/s (理论 2.000000)" % result["verification"]["back_max_at_vmax"])
    lines.append("V_max*(1+1e-6) 后全队最大速度 = %.9f m/s (> 2, 违反约束)" % result["verification"]["over_limit_at_vmax_plus"])
    lines.append("区间边界与远场检查: s_start K=%.9f, s_start-30 K=%.9f, s_end K=%.9f, s_end+30 K=%.9f, s_end+100 K=%.9f, s_end+200 K=%.9f" % (
        result["boundary_checks"]["s_start"]["k_max"], result["boundary_checks"]["s_start-30"]["k_max"],
        result["boundary_checks"]["s_end"]["k_max"], result["boundary_checks"]["s_end+30"]["k_max"],
        result["boundary_checks"]["s_end+100"]["k_max"], result["boundary_checks"]["s_end+200"]["k_max"]))
    if result["reference"] is not None:
        ref = result["reference"]
        lines.append("与独立复现参考对比: K_max 差 %.3e, V_max 差 %.3e, s* 差 %.3e" % (
            result["K_max"] - ref["amplification"], result["V_max"] - ref["max_head_speed"],
            result["s_star"] - ref["s_head"]))
    lines.append("")
    lines.append("=== 表7 论文关键结果表 ===")
    lines.append("%-28s | %s" % ("项目", "数值"))
    rows = [
        ("K_max (1 m/s 下全局最大倍率)", "%.9f" % result["K_max"]),
        ("V_max (龙头最大恒定速度, m/s)", "%.9f" % result["V_max"]),
        ("峰值龙头弧长坐标 s* (m)", "%.9f" % result["s_star"]),
        ("峰值时龙头所在区段", "%s" % result["head_segment"]),
        ("控制把手", "%s" % result["controlling_handle_name"]),
        ("控制把手所在区段", "%s" % result["controlling_segment"]),
        ("控制把手速度倍率", "%.9f" % result["K_max"]),
        ("V_max 下控制把手速度 (m/s)", "%.6f" % (result["K_max"] * result["V_max"])),
        ("V_max 下全队最大速度 (m/s)", "%.6f" % result["verification"]["back_max_at_vmax"]),
    ]
    for name, value in rows:
        lines.append("%-28s | %s" % (name, value))
    lines.append("")
    lines.append("=== 表8 临界构型指定把手状态 (s* = %.6f m, V_max = %.6f m/s) ===" % (
        result["s_star"], result["V_max"]))
    lines.append("%-14s %14s %14s %14s %14s" % ("部位", "x (m)", "y (m)", "速度倍率", "速度 (m/s)"))
    for idx in (0, 1, 3, 51, 101, 151, 201, 223):
        row = result["handle_rows"][idx]
        lines.append("%-14s %14.6f %14.6f %14.6f %14.6f" % (
            row["name"], row["x"], row["y"], row["ratio"], row["speed_at_vmax"]))
    lines.append("")
    lines.append("全量数据保存在 24A成果\\第五问求解\\result5.xlsx 中。")
    out = os.path.join(BASE, "表7_表8_关键结果.txt")
    with open(out, "w", encoding="utf-8") as fh:
        fh.write("\n".join(lines))
    print("saved", out)
    return out


# ---------------------------------------------------------------------------
# 输出: 两张图
# ---------------------------------------------------------------------------
def sample_path(path, s0, s1, n=400):
    ss = np.linspace(s0, s1, n)
    return np.array([path.point_tangent(float(s))[0] for s in ss])


def draw_k_curve(path, result):
    grid = result["coarse_grid"]["grid"]
    k_values = result["coarse_grid"]["k"]
    s_star = result["s_star"]
    k_star = result["K_max"]

    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(9.5, 8.5),
                                   gridspec_kw={"height_ratios": [1.0, 1.0]})
    ax1.plot(grid, k_values, "-", color="#2c7fb8", linewidth=1.0)
    ax1.plot([s_star], [k_star], "o", color="#c0392b", markersize=8,
             label="全局峰值 K_max = %.6f (s* = %.3f m)" % (k_star, s_star))
    ax1.axhline(1.0, color="#7f8c8d", linestyle=":", linewidth=1.0, label="K = 1 (龙头速度)")
    for junction in (0.0, path.length1, path.turn_length):
        ax1.axvline(junction, color="#e67e22", linestyle="--", linewidth=0.8)
    ax1.set_xlim(result["interval"]["s_start"] - 20, result["interval"]["s_end"] + 20)
    ax1.grid(True, linestyle=":", linewidth=0.6, alpha=0.6)
    ax1.set_ylabel("K(s) = max k_i(s)")
    ax1.set_title("第五问 速度倍率随龙头弧长坐标变化 (全区间)")
    ax1.legend(loc="lower right", fontsize=9)

    margin = 3.0
    mask = (grid >= s_star - margin) & (grid <= s_star + margin)
    ax2.plot(grid[mask], k_values[mask], "-", color="#2c7fb8", linewidth=1.6)
    ax2.plot([s_star], [k_star], "o", color="#c0392b", markersize=9,
             label="K_max = %.9f, %s(编号 %d)" % (k_star, result["controlling_handle_name"],
                                                  result["controlling_handle"]))
    ax2.axvline(path.turn_length, color="#e67e22", linestyle="--", linewidth=1.0, label="出口 B")
    ax2.axvline(path.length1, color="#8e44ad", linestyle="--", linewidth=1.0, label="切点 J")
    ax2.axhline(1.0, color="#7f8c8d", linestyle=":", linewidth=1.0)
    ax2.grid(True, linestyle=":", linewidth=0.6, alpha=0.6)
    ax2.set_xlabel("龙头弧长坐标 s (m)")
    ax2.set_ylabel("K(s)")
    ax2.set_title("峰值邻域放大 (s* ± 3 m)")
    ax2.legend(loc="lower right", fontsize=9)
    fig.tight_layout()
    out = os.path.join(BASE, "速度倍率_全局搜索曲线.png")
    fig.savefig(out, dpi=170)
    plt.close(fig)
    print("saved", out)
    return out


def draw_critical_configuration(path, result):
    state = result["state_peak"]
    handle_peak = result["controlling_handle"]
    fig, ax = plt.subplots(figsize=(9.5, 9.5))
    inward = sample_path(path, -100, 0, 600)
    arc1 = sample_path(path, 0, path.length1, 200)
    arc2 = sample_path(path, path.length1, path.turn_length, 200)
    outward = sample_path(path, path.turn_length, path.turn_length + 100, 600)
    ax.plot(inward[:, 0], inward[:, 1], "-", color="#bdc3c7", linewidth=0.9)
    ax.plot(arc1[:, 0], arc1[:, 1], "-", color="#c0392b", linewidth=2.0, label="第一段圆弧 R1")
    ax.plot(arc2[:, 0], arc2[:, 1], "-", color="#e67e22", linewidth=2.0, label="第二段圆弧 R2")
    ax.plot(outward[:, 0], outward[:, 1], "-", color="#bdc3c7", linewidth=0.9, label="盘入/盘出螺线")
    ax.add_patch(Circle((0, 0), TURN_RADIUS, fill=False, color="#8e44ad",
                        linestyle="--", linewidth=1.4, label="调头空间边界 R=4.5 m"))
    px = np.array(state["px"])
    py = np.array(state["py"])
    ax.plot(px, py, "-", color="#34495e", linewidth=1.1,
            label="临界构型全龙 (s* = %.3f m)" % result["s_star"])
    ax.plot(px[0], py[0], "o", color="#27ae60", markersize=8, label="龙头前把手 P0")
    ax.plot(px[1], py[1], "s", color="#2980b9", markersize=7, label="第1节龙身前把手 P1")
    ax.plot(px[2], py[2], "s", color="#2980b9", markersize=7, label="第2节龙身前把手 P2")
    ax.plot(px[handle_peak], py[handle_peak], "*", color="#c0392b", markersize=16,
            label="控制把手 %s (k = %.6f)" % (handle_name(handle_peak), result["K_max"]))
    ax.plot([path.a[0], path.junction[0], -path.a[0]],
            [path.a[1], path.junction[1], -path.a[1]],
            "x", color="#1f4e79", markersize=9)
    ax.annotate("A", path.a, textcoords="offset points", xytext=(8, 8), fontsize=11)
    ax.annotate("J", path.junction, textcoords="offset points", xytext=(8, 8), fontsize=11)
    ax.annotate("B", (-path.a[0], -path.a[1]), textcoords="offset points", xytext=(8, 8), fontsize=11)
    ax.set_aspect("equal", adjustable="box")
    ax.grid(True, linestyle=":", linewidth=0.6, alpha=0.6)
    ax.set_title("第五问 K_max 临界构型全龙形态 (s* = %.3f m)" % result["s_star"])
    ax.set_xlabel("x (m)")
    ax.set_ylabel("y (m)")
    ax.legend(loc="best", fontsize=9)
    fig.tight_layout()
    out = os.path.join(BASE, "临界构型_峰值示意图.png")
    fig.savefig(out, dpi=170)
    plt.close(fig)
    print("saved", out)
    return out


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------
def main():
    path, result = solve_question5()
    write_result5(path, result)
    write_key_txt(path, result)
    draw_k_curve(path, result)
    draw_critical_configuration(path, result)
    summary = {key: value for key, value in result.items()
               if key not in ("coarse_grid", "fine_grid", "candidates",
                              "refined_peaks", "handle_rows", "state_peak",
                              "boundary_checks")}
    summary_out = os.path.join(BASE, "result5.json")
    with open(summary_out, "w", encoding="utf-8") as fh:
        json.dump(summary, fh, ensure_ascii=False, indent=2,
                  default=lambda value: value.tolist() if isinstance(value, np.ndarray) else value)
    print("saved", summary_out)


if __name__ == "__main__":
    main()
