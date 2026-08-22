#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""核验 Q3 JSON、官方 Excel 模板输出及物理约束。"""

from __future__ import annotations

import json
import math
from pathlib import Path
import sys

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
JSON_PATH = HERE / "q3_runs" / "q3_best.json"
XLSX_PATH = HERE / "result1.xlsx"

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")


def close(a: float, b: float, tol: float = 1e-5) -> bool:
    return math.isclose(float(a), float(b), rel_tol=0.0, abs_tol=tol)


def main() -> None:
    failures: list[str] = []
    if not JSON_PATH.exists():
        failures.append(f"缺少 {JSON_PATH}")
    if not XLSX_PATH.exists() or XLSX_PATH.stat().st_size < 1000:
        failures.append(f"缺少或文件过小：{XLSX_PATH}")
    if failures:
        raise SystemExit("\n".join(failures))

    result = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    bombs = result["bombs"]
    if len(bombs) != 3:
        failures.append("JSON 中烟幕弹数量不是 3")
    for i in range(2):
        if bombs[i + 1]["drop_time"] - bombs[i]["drop_time"] < 1.0 - 1e-9:
            failures.append(f"弹{i + 1}与弹{i + 2}投放间隔小于 1 s")
    if not 70.0 <= result["speed"] <= 140.0:
        failures.append("无人机速度越界")
    if any(b["explosion_point"][2] < -1e-9 for b in bombs):
        failures.append("存在地下起爆点")

    workbook = load_workbook(XLSX_PATH, data_only=False)
    sheet = workbook.active
    for row, bomb in enumerate(bombs, 2):
        expected = [
            result["heading_deg"], result["speed"], bomb["index"],
            *bomb["drop_point"], *bomb["explosion_point"], bomb["duration"],
        ]
        actual = [sheet.cell(row=row, column=column).value for column in range(1, 11)]
        for column, (got, want) in enumerate(zip(actual, expected), 1):
            if got is None or not close(got, want):
                failures.append(f"Excel {sheet.cell(row=row, column=column).coordinate} 与 JSON 不一致")
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("#"):
                failures.append(f"Excel 错误值：{cell.coordinate}={cell.value}")

    if failures:
        print("[FAIL] Q3 产物核验失败")
        for failure in failures:
            print(" -", failure)
        raise SystemExit(1)
    print("[OK] 三枚弹、投放间隔、速度、起爆高度均满足约束")
    print("[OK] result1.xlsx 与 q3_best.json 的 30 个结果单元格一致")
    print("[OK] 工作簿无 Excel 错误值")
    print(f"[OK] 严格并集时长 {result['independent_union_duration']:.10f} s")


if __name__ == "__main__":
    main()
