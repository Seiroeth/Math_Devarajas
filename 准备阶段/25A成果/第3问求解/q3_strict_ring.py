#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""2025 A 题问题 3：FY1 三枚烟幕弹严格圆柱遮蔽优化。"""

from __future__ import annotations

import argparse
from datetime import datetime
import json
import math
from pathlib import Path
import shutil
import sys
import time

import matplotlib.pyplot as plt
from matplotlib import font_manager
import numpy as np
from openpyxl import load_workbook
from scipy.optimize import differential_evolution, minimize

from smoke_strict_core import (
    G,
    MISSILE_HIT_TIME,
    Smoke,
    cooperative_intervals,
    independent_union,
    interval_length,
)

HERE = Path(__file__).resolve().parent


def find_project_root(start: Path) -> Path:
    """允许脚本位于 25A成果 根目录或其题目子目录。"""
    for candidate in (start, *start.parents):
        if (candidate / "数学建模学习资料").is_dir() and (candidate / "准备阶段").is_dir():
            return candidate
    raise FileNotFoundError("无法定位 Math_Devarajas 项目根目录。")


PROJECT_ROOT = find_project_root(HERE)
TEMPLATE = PROJECT_ROOT / "数学建模学习资料" / "历年赛题题目" / "2025A" / "附件" / "result1.xlsx"
RUN_DIR = HERE / "q3_runs"
LOG_DIR = RUN_DIR / "logs"
CHECKPOINT_DIR = RUN_DIR / "checkpoints"
RESULT_JSON = RUN_DIR / "q3_best.json"
RESULT_XLSX = HERE / "result1.xlsx"

# 网页思路中引用的开源方案，仅作为起点和数量级 benchmark；最终由本模型严格复算。
REFERENCE = np.array([
    179.6475,
    139.9983,
    0.0030,
    3.7025 - 0.0030 - 1.0,
    5.5695 - 3.7025 - 1.0,
    3.6111,
    5.3375,
    6.0405,
])

# 在本项目连续圆周判据下，把三段区间端点对齐并逐坐标精修后的候选。
REFINED = np.array([
    179.6500000000,
    140.0000000000,
    0.0000000000,
    2.7167561586,
    0.8718282145,
    3.6050000000,
    5.3400000000,
    6.0500000000,
])

BOUNDS = [
    (150.0, 210.0),
    (70.0, 140.0),
    (0.0, 8.0),
    (0.0, 10.0),
    (0.0, 10.0),
    (0.05, math.sqrt(2.0 * 1800.0 / G)),
    (0.05, math.sqrt(2.0 * 1800.0 / G)),
    (0.05, math.sqrt(2.0 * 1800.0 / G)),
]

SEARCH = {
    "quick": {"dt": 0.08, "nphi": 72, "maxiter": 35, "popsize": 8, "tol": 2e-5},
    "full": {"dt": 0.04, "nphi": 180, "maxiter": 140, "popsize": 16, "tol": 2e-7},
}

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


def decode(x: np.ndarray) -> list[Smoke]:
    psi_deg, speed, td1, s2, s3, tau1, tau2, tau3 = map(float, x)
    td2 = td1 + 1.0 + s2
    td3 = td2 + 1.0 + s3
    psi = math.radians(psi_deg % 360.0)
    return [
        Smoke(psi, speed, td1, tau1),
        Smoke(psi, speed, td2, tau2),
        Smoke(psi, speed, td3, tau3),
    ]


def encode(smokes: list[Smoke]) -> np.ndarray:
    return np.array([
        math.degrees(smokes[0].psi) % 360.0,
        smokes[0].speed,
        smokes[0].td,
        smokes[1].td - smokes[0].td - 1.0,
        smokes[2].td - smokes[1].td - 1.0,
        smokes[0].tau,
        smokes[1].tau,
        smokes[2].tau,
    ])


def write_json(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_suffix(path.suffix + ".tmp")
    temp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    temp.replace(path)


def load_resume_x() -> np.ndarray:
    candidates = [REFERENCE, REFINED]
    for path in [RESULT_JSON, *sorted(CHECKPOINT_DIR.glob("*.json"))]:
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            x = data.get("x") or data.get("best_x")
            if x and len(x) == 8:
                candidates.append(np.asarray(x, dtype=float))
        except (OSError, ValueError, TypeError):
            continue
    return max(candidates, key=lambda x: fast_duration(x, dt=0.02, nphi=360))


def fast_duration(x: np.ndarray, dt: float, nphi: int) -> float:
    smokes = decode(np.asarray(x, dtype=float))
    if not all(smoke.valid for smoke in smokes):
        return 0.0
    if min(smokes[i + 1].td - smokes[i].td for i in range(2)) < 1.0 - 1e-10:
        return 0.0
    return independent_union(smokes, exact=False, dt=dt, nphi=nphi)[2]


class SearchTracker:
    def __init__(self, mode: str, dt: float, nphi: int, start_x: np.ndarray):
        self.mode = mode
        self.dt = dt
        self.nphi = nphi
        self.generation = 0
        self.started = time.time()
        self.best_x = np.asarray(start_x, dtype=float)
        self.best_duration = fast_duration(self.best_x, dt, nphi)
        self.status_path = LOG_DIR / "status.json"

    def objective(self, x: np.ndarray) -> float:
        duration = fast_duration(x, self.dt, self.nphi)
        if duration > self.best_duration:
            self.best_duration = duration
            self.best_x = np.asarray(x, dtype=float).copy()
        return -duration

    def callback(self, xk: np.ndarray, convergence: float) -> bool:
        self.generation += 1
        duration = fast_duration(xk, self.dt, self.nphi)
        if duration > self.best_duration:
            self.best_duration = duration
            self.best_x = np.asarray(xk, dtype=float).copy()
        status = {
            "mode": self.mode,
            "generation": self.generation,
            "best_duration_fast": self.best_duration,
            "best_x": self.best_x.tolist(),
            "convergence": float(convergence),
            "elapsed_seconds": time.time() - self.started,
            "updated_at": datetime.now().isoformat(timespec="seconds"),
        }
        write_json(self.status_path, status)
        write_json(CHECKPOINT_DIR / f"{self.mode}_latest.json", status)
        print(
            f"generation={self.generation:03d}  fast={self.best_duration:.6f} s  "
            f"conv={convergence:.3e}",
            flush=True,
        )
        return False


def optimize(mode: str, force: bool = False) -> np.ndarray:
    settings = SEARCH[mode]
    checkpoint = CHECKPOINT_DIR / f"{mode}_complete.json"
    if checkpoint.exists() and not force:
        data = json.loads(checkpoint.read_text(encoding="utf-8"))
        print(f"读取已完成的 {mode} 检查点；如需重跑请加 --force。")
        saved = np.asarray(data["x"], dtype=float)
        return max([saved, REFINED], key=lambda x: fast_duration(x, dt=0.02, nphi=360))

    start_x = load_resume_x()
    tracker = SearchTracker(mode, settings["dt"], settings["nphi"], start_x)
    result = differential_evolution(
        tracker.objective,
        bounds=BOUNDS,
        seed=2025,
        popsize=settings["popsize"],
        maxiter=settings["maxiter"],
        tol=settings["tol"],
        polish=False,
        updating="immediate",
        workers=1,
        x0=start_x,
        callback=tracker.callback,
    )
    candidates = [REFERENCE, REFINED, start_x, tracker.best_x, np.asarray(result.x)]

    local = minimize(
        tracker.objective,
        min(candidates, key=tracker.objective),
        method="Powell",
        bounds=BOUNDS,
        options={"maxiter": 220 if mode == "full" else 90, "xtol": 1e-5, "ftol": 1e-6},
    )
    candidates.append(np.asarray(local.x))
    best = max(candidates, key=lambda x: fast_duration(x, dt=0.02, nphi=360))
    data = {
        "mode": mode,
        "x": best.tolist(),
        "fast_duration_fine": fast_duration(best, dt=0.02, nphi=360),
        "completed_at": datetime.now().isoformat(timespec="seconds"),
    }
    write_json(checkpoint, data)
    return best


def exact_report(x: np.ndarray, cooperative: bool = True) -> dict:
    smokes = decode(x)
    each, union, total = independent_union(smokes, exact=True)
    result = {
        "x": np.asarray(x, dtype=float).tolist(),
        "heading_deg": math.degrees(smokes[0].psi) % 360.0,
        "speed": smokes[0].speed,
        "missile_hit_time": MISSILE_HIT_TIME,
        "bombs": [],
        "independent_union_intervals": [[a, b] for a, b in union],
        "independent_union_duration": total,
    }
    for index, (smoke, intervals) in enumerate(zip(smokes, each), 1):
        result["bombs"].append({
            "index": index,
            "drop_time": smoke.td,
            "fuse_delay": smoke.tau,
            "explosion_time": smoke.te,
            "drop_point": smoke.drop_point.tolist(),
            "explosion_point": smoke.explosion_point.tolist(),
            "intervals": [[a, b] for a, b in intervals],
            "duration": interval_length(intervals),
        })
    if cooperative:
        coop, point_count = cooperative_intervals(
            smokes,
            time_dt=0.02,
            theta_step_deg=1.0,
            z_step=0.5,
            radial_step=0.5,
        )
        coop_total = interval_length(coop)
        result.update({
            "cooperative_surface_points": point_count,
            "cooperative_intervals": [[a, b] for a, b in coop],
            "cooperative_duration": coop_total,
            "cooperative_gain": coop_total - total,
        })
    result["verified_at"] = datetime.now().isoformat(timespec="seconds")
    return result


def export_xlsx(result: dict, path: Path = RESULT_XLSX) -> None:
    if not TEMPLATE.exists():
        raise FileNotFoundError(f"未找到官方模板：{TEMPLATE}")
    shutil.copy2(TEMPLATE, path)
    workbook = load_workbook(path)
    sheet = workbook.active
    for row, bomb in enumerate(result["bombs"], 2):
        values = [
            result["heading_deg"],
            result["speed"],
            bomb["index"],
            *bomb["drop_point"],
            *bomb["explosion_point"],
            bomb["duration"],
        ]
        for column, value in enumerate(values, 1):
            sheet.cell(row=row, column=column, value=float(value) if column != 3 else int(value))
        for column in range(1, 11):
            sheet.cell(row=row, column=column).number_format = "0.000000"
        sheet.cell(row=row, column=3).number_format = "0"
    sheet.print_area = "A1:J6"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.35
    sheet.page_margins.bottom = 0.35
    workbook.save(path)


def plot_timeline(result: dict, path: Path) -> None:
    font_path = font_manager.findfont("Microsoft YaHei")
    plt.rcParams["font.family"] = font_manager.FontProperties(fname=font_path).get_name()
    plt.rcParams["axes.unicode_minus"] = False
    fig, ax = plt.subplots(figsize=(10, 3.8))
    colors = ["#4472C4", "#ED7D31", "#70AD47"]
    for i, bomb in enumerate(result["bombs"]):
        for a, b in bomb["intervals"]:
            ax.barh(i + 1, b - a, left=a, height=0.55, color=colors[i])
    for a, b in result["independent_union_intervals"]:
        ax.barh(0, b - a, left=a, height=0.55, color="#7030A0")
    ax.set_yticks([0, 1, 2, 3], ["并集", "烟幕弹1", "烟幕弹2", "烟幕弹3"])
    ax.set_xlabel("任务开始后的时间 / s")
    ax.set_title(f"Q3 严格圆柱遮蔽时间轴（并集 {result['independent_union_duration']:.6f} s）")
    ax.grid(axis="x", alpha=0.25)
    fig.tight_layout()
    fig.savefig(path, dpi=180)
    plt.close(fig)


def print_report(result: dict) -> None:
    print("=== Q3 严格圆柱判据 ===")
    print(f"航向角: {result['heading_deg']:.9f} deg")
    print(f"速度  : {result['speed']:.9f} m/s")
    for bomb in result["bombs"]:
        print(
            f"弹{bomb['index']}: td={bomb['drop_time']:.9f}, tau={bomb['fuse_delay']:.9f}, "
            f"te={bomb['explosion_time']:.9f}, duration={bomb['duration']:.9f}"
        )
        print("  投放点:", np.array2string(np.asarray(bomb["drop_point"]), precision=6))
        print("  起爆点:", np.array2string(np.asarray(bomb["explosion_point"]), precision=6))
        print("  区间  :", bomb["intervals"])
    print("独立完整遮蔽并集:", result["independent_union_intervals"])
    print(f"独立完整遮蔽总时长: {result['independent_union_duration']:.10f} s")
    if "cooperative_duration" in result:
        print(f"表面 max-min 联合时长: {result['cooperative_duration']:.10f} s")
        print(f"协同增益: {result['cooperative_gain']:.10f} s")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--mode", choices=["verify", "quick", "full"], default="verify")
    parser.add_argument("--force", action="store_true", help="忽略已完成检查点并重跑搜索")
    parser.add_argument("--skip-coop", action="store_true", help="跳过耗时的多烟幕表面精检")
    args = parser.parse_args()

    LOG_DIR.mkdir(parents=True, exist_ok=True)
    CHECKPOINT_DIR.mkdir(parents=True, exist_ok=True)
    x = REFINED.copy() if args.mode == "verify" else optimize(args.mode, force=args.force)
    result = exact_report(x, cooperative=not args.skip_coop)
    write_json(RESULT_JSON, result)
    export_xlsx(result)
    plot_timeline(result, HERE / "q3_timeline.png")
    print_report(result)
    print(f"JSON : {RESULT_JSON}")
    print(f"Excel: {RESULT_XLSX}")


if __name__ == "__main__":
    main()
