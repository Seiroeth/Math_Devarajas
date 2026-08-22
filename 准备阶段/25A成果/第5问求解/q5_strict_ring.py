#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""2025 国赛 A 题第 5 问：五机、至多三弹、三导弹严格圆柱全遮蔽。

默认 ``python q5_strict_ring.py`` 会从已验证的精修策略出发，用连续圆周极值
和 Brent 根求解终算，并生成 result3.xlsx。也可分别运行：

* ``--mode verify``：不搜索，约数秒复算 warm-start 并生成全部结果；
* ``--mode optimize``：断点式局部精修（默认模式）；
* ``--mode probe``：独立搜索 15 个 UAV–导弹配对，每一对单独保存 JSON。

角度约定：从 +x 轴逆时针，输出到 [0, 360) 度。
"""

from __future__ import annotations

import argparse
from dataclasses import asdict, dataclass, replace
from datetime import datetime
import importlib.util
import json
import math
from pathlib import Path
import shutil
import sys
import time
from typing import Iterable

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from scipy.optimize import differential_evolution


HERE = Path(__file__).resolve().parent
OUTPUT_ROOT = HERE.parent
RUN_DIR = HERE / "q5_runs"
PROBE_DIR = RUN_DIR / "probes"
BEST_JSON = RUN_DIR / "q5_best.json"
STATUS_JSON = RUN_DIR / "status.json"
LOG_FILE = RUN_DIR / "q5_run.log"
RESULT_XLSX = HERE / "result3.xlsx"
TIMELINE_PNG = HERE / "q5_timeline.png"


def find_project_root() -> Path:
    """找到同时包含学习资料与准备阶段的 Math_Devarajas 根目录。"""
    for candidate in (HERE, *HERE.parents):
        if (
            (candidate / "数学建模学习资料").is_dir()
            and (candidate / "准备阶段").is_dir()
        ):
            return candidate
    return OUTPUT_ROOT


PROJECT_ROOT = find_project_root()
TEMPLATE_XLSX = (
    PROJECT_ROOT
    / "数学建模学习资料"
    / "历年赛题题目"
    / "2025A"
    / "附件"
    / "result3.xlsx"
)


def load_core():
    """不依赖当前工作目录加载第 3 问中的公共严格几何核心。"""
    candidates = [
        HERE / "smoke_strict_core.py",
        OUTPUT_ROOT / "smoke_strict_core.py",
        OUTPUT_ROOT / "第3问求解" / "smoke_strict_core.py",
        OUTPUT_ROOT / "q3_runs" / "第3问求解过程文件" / "smoke_strict_core.py",
    ]
    for path in candidates:
        if path.exists():
            spec = importlib.util.spec_from_file_location("smoke_strict_core", path)
            if spec is None or spec.loader is None:
                continue
            module = importlib.util.module_from_spec(spec)
            sys.modules[spec.name] = module
            spec.loader.exec_module(module)
            return module
    searched = "\n".join(str(path) for path in candidates)
    raise FileNotFoundError(f"找不到 smoke_strict_core.py，已检查：\n{searched}")


core = load_core()
MISSILES = core.MISSILES
Smoke = core.Smoke
UAV_POSITIONS = core.UAV_POSITIONS
cooperative_intervals = core.cooperative_intervals
fast_intervals = core.fast_intervals
independent_union = core.independent_union
interval_length = core.interval_length

UAV_NAMES = tuple(UAV_POSITIONS)
MISSILE_NAMES = tuple(MISSILES)


@dataclass(frozen=True)
class BombPlan:
    uav: str
    missile: str
    heading_deg: float
    speed: float
    drop_time: float
    fuse_delay: float
    bomb_id: int

    def smoke(self) -> Smoke:
        return Smoke(
            psi=math.radians(self.heading_deg % 360.0),
            speed=self.speed,
            td=self.drop_time,
            tau=self.fuse_delay,
            u0=UAV_POSITIONS[self.uav],
        )


# 网页思路给出的分派作为 warm-start；FY1 使用本目录第 3 问已进一步精修的方案。
WARM_START = [
    BombPlan("FY1", "M1", 179.65, 140.0, 0.0, 3.605, 1),
    BombPlan("FY1", "M1", 179.65, 140.0, 3.7167561586, 5.34, 2),
    BombPlan("FY1", "M1", 179.65, 140.0, 5.5885843731, 6.05, 3),
    # 在网页种子 (293.6618°, 140 m/s) 上按本程序严格判据再次精修。
    BombPlan("FY2", "M2", 293.8797956, 139.99421745, 5.50500146, 2.00422709, 1),
    BombPlan("FY2", "M2", 293.8797956, 139.99421745, 6.65043742, 0.97895505, 2),
    BombPlan("FY3", "M2", 86.80310596231972, 133.4, 24.29, 0.3, 1),
    BombPlan("FY4", "M2", 237.6055976407524, 81.6, 12.54, 11.86, 1),
    BombPlan("FY5", "M3", 116.37753910706417, 140.0, 11.808152214609857, 1.0403723078137124, 1),
]

# 单弹探针的定向种子；无种子组合自动使用指向假目标、真目标和导弹初始点的方向。
PAIR_SEEDS = {
    ("FY1", "M1"): WARM_START[0],
    ("FY2", "M1"): BombPlan("FY2", "M1", 306.1909, 136.85, 8.551, 3.986, 1),
    ("FY2", "M2"): WARM_START[3],
    ("FY3", "M1"): BombPlan("FY3", "M1", 122.4743, 92.96, 31.733, 7.646, 1),
    ("FY3", "M2"): WARM_START[5],
    ("FY4", "M2"): WARM_START[6],
    ("FY5", "M1"): BombPlan("FY5", "M1", 102.0, 136.7, 12.55, 2.64, 1),
    ("FY5", "M3"): WARM_START[7],
}


def json_ready(value):
    if isinstance(value, np.ndarray):
        return value.tolist()
    if isinstance(value, (np.floating, np.integer)):
        return value.item()
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, dict):
        return {str(k): json_ready(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [json_ready(v) for v in value]
    return value


def atomic_write_json(path: Path, value) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_suffix(path.suffix + ".tmp")
    temp.write_text(
        json.dumps(json_ready(value), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    temp.replace(path)


def log(message: str) -> None:
    RUN_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().isoformat(timespec="seconds")
    line = f"[{stamp}] {message}"
    print(line, flush=True)
    with LOG_FILE.open("a", encoding="utf-8") as stream:
        stream.write(line + "\n")


def validate_plan(plans: Iterable[BombPlan]) -> list[str]:
    plans = list(plans)
    errors: list[str] = []
    by_uav = {name: [] for name in UAV_NAMES}
    for plan in plans:
        if plan.uav not in UAV_POSITIONS:
            errors.append(f"未知无人机 {plan.uav}")
            continue
        if plan.missile not in MISSILES:
            errors.append(f"未知导弹 {plan.missile}")
        if not 70.0 <= plan.speed <= 140.0:
            errors.append(f"{plan.uav}-{plan.bomb_id} 速度越界")
        if plan.drop_time < 0.0 or plan.fuse_delay < 0.0:
            errors.append(f"{plan.uav}-{plan.bomb_id} 时间为负")
        smoke = plan.smoke()
        if smoke.explosion_point[2] < -1e-9:
            errors.append(f"{plan.uav}-{plan.bomb_id} 在地下起爆")
        if smoke.te >= MISSILES[plan.missile].hit_time:
            errors.append(f"{plan.uav}-{plan.bomb_id} 起爆晚于 {plan.missile} 命中")
        by_uav[plan.uav].append(plan)

    for uav, bombs in by_uav.items():
        if len(bombs) > 3:
            errors.append(f"{uav} 投弹 {len(bombs)} 枚，超过 3 枚")
        if not bombs:
            continue
        h0, v0 = bombs[0].heading_deg % 360.0, bombs[0].speed
        for bomb in bombs[1:]:
            angle_gap = abs((bomb.heading_deg - h0 + 180.0) % 360.0 - 180.0)
            if angle_gap > 1e-8 or abs(bomb.speed - v0) > 1e-8:
                errors.append(f"{uav} 的多枚弹没有共享航向/速度")
        ordered = sorted(bombs, key=lambda b: b.drop_time)
        for left, right in zip(ordered[:-1], ordered[1:]):
            if right.drop_time - left.drop_time < 1.0 - 1e-9:
                errors.append(f"{uav} 相邻投弹间隔小于 1 s")
    return errors


def plan_records(plans: list[BombPlan], per_missile) -> list[dict]:
    records = []
    for index, plan in enumerate(plans):
        smoke = plan.smoke()
        intervals = per_missile[plan.missile][index]
        record = asdict(plan)
        record.update(
            explosion_time=smoke.te,
            drop_point=smoke.drop_point.tolist(),
            explosion_point=smoke.explosion_point.tolist(),
            assigned_intervals=[[a, b] for a, b in intervals],
            assigned_duration=interval_length(intervals),
        )
        records.append(record)
    return records


def evaluate_plan(
    plans: list[BombPlan],
    *,
    exact: bool,
    cooperative: bool = False,
    fine_cooperative: bool = False,
) -> dict:
    errors = validate_plan(plans)
    if errors:
        raise ValueError("策略不可行：\n- " + "\n- ".join(errors))
    smokes = [plan.smoke() for plan in plans]
    per_missile = {}
    missile_results = {}
    for name in MISSILE_NAMES:
        each, union, total = independent_union(
            smokes,
            exact=exact,
            dt=0.04 if not exact else 0.08,
            nphi=144 if not exact else 72,
            missile=MISSILES[name],
        )
        per_missile[name] = each
        missile_results[name] = {
            "hit_time": MISSILES[name].hit_time,
            "independent_intervals": [[a, b] for a, b in union],
            "independent_duration": total,
        }

    result = {
        "mode": "strict_continuous_target_rings",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "objective_definition": "sum of union full-cover durations for M1, M2 and M3",
        "missiles": missile_results,
        "independent_total": float(
            sum(item["independent_duration"] for item in missile_results.values())
        ),
        "bombs": plan_records(plans, per_missile),
        "constraints": {
            "valid": True,
            "max_bombs_per_uav": 3,
            "minimum_drop_gap": 1.0,
            "same_heading_speed_per_uav": True,
        },
    }

    if cooperative:
        coop_total = 0.0
        for name in MISSILE_NAMES:
            if fine_cooperative:
                settings = dict(
                    time_dt=0.02,
                    theta_step_deg=1.0,
                    z_step=0.5,
                    radial_step=0.5,
                )
            else:
                settings = dict(
                    time_dt=0.03,
                    theta_step_deg=2.0,
                    z_step=1.0,
                    radial_step=1.0,
                )
            intervals, point_count = cooperative_intervals(
                smokes, missile=MISSILES[name], **settings,
            )
            duration = interval_length(intervals)
            result["missiles"][name].update(
                cooperative_surface_points=point_count,
                cooperative_intervals=[[a, b] for a, b in intervals],
                cooperative_duration=duration,
                cooperative_gain=(
                    duration
                    - result["missiles"][name]["independent_duration"]
                ),
            )
            coop_total += duration
        result["cooperative_total"] = coop_total
    return result


def fast_score(plans: list[BombPlan], missile_name: str) -> float:
    if validate_plan(plans):
        return -1e6
    smokes = [plan.smoke() for plan in plans]
    return independent_union(
        smokes,
        exact=False,
        dt=0.055,
        nphi=96,
        missile=MISSILES[missile_name],
    )[2]


def aim_headings(uav: str, missile_name: str) -> list[float]:
    p = np.asarray(UAV_POSITIONS[uav], dtype=float)
    target_xy = [np.array([0.0, 0.0]), np.array([0.0, 200.0])]
    target_xy.append(MISSILES[missile_name].initial_position[:2])
    return [
        math.degrees(math.atan2(q[1] - p[1], q[0] - p[0])) % 360.0
        for q in target_xy
    ]


def probe_pair(
    uav: str,
    missile_name: str,
    iterations: int,
    seed: int,
    force: bool,
) -> dict:
    """单弹宽域探针；完成一对立即写盘，可安全断点续跑。"""
    output = PROBE_DIR / f"{uav}_{missile_name}.json"
    if output.exists() and not force:
        return json.loads(output.read_text(encoding="utf-8"))
    missile = MISSILES[missile_name]
    z0 = UAV_POSITIONS[uav][2]
    max_tau = min(16.0, math.sqrt(2.0 * z0 / core.G))
    max_td = min(50.0, missile.hit_time - 0.31)
    bounds = [(0.0, 360.0), (70.0, 140.0), (0.0, max_td), (0.3, max_tau)]

    def objective(x) -> float:
        plan = BombPlan(uav, missile_name, *map(float, x), bomb_id=1)
        return -fast_score([plan], missile_name)

    known = PAIR_SEEDS.get((uav, missile_name))
    if known is not None:
        x0 = np.array([
            known.heading_deg,
            known.speed,
            known.drop_time,
            known.fuse_delay,
        ])
    else:
        headings = aim_headings(uav, missile_name)
        x0 = np.array([headings[1], 120.0, min(5.0, max_td), min(3.0, max_tau)])
    lower = np.asarray([item[0] for item in bounds], dtype=float)
    upper = np.asarray([item[1] for item in bounds], dtype=float)
    epsilon = 1e-10 * np.maximum(1.0, upper - lower)
    x0 = np.minimum(np.maximum(x0, lower + epsilon), upper - epsilon)
    result = differential_evolution(
        objective,
        bounds,
        seed=seed,
        maxiter=max(1, iterations),
        popsize=7,
        tol=1e-6,
        polish=True,
        updating="immediate",
        workers=1,
        x0=x0,
    )
    candidates = [result.x, x0]
    # 把三种物理指向也纳入候选，减少低预算探针漏检。
    for heading in aim_headings(uav, missile_name):
        candidates.append(np.array([heading, 120.0, min(5.0, max_td), min(3.0, max_tau)]))
    best_x = max(candidates, key=lambda x: -objective(x))
    best_plan = BombPlan(uav, missile_name, *map(float, best_x), bomb_id=1)
    exact_intervals = core.exact_intervals(best_plan.smoke(), missile=missile)
    payload = {
        "uav": uav,
        "missile": missile_name,
        "parameters": asdict(best_plan),
        "fast_duration": fast_score([best_plan], missile_name),
        "exact_intervals": [[a, b] for a, b in exact_intervals],
        "exact_duration": interval_length(exact_intervals),
        "optimizer_message": str(result.message),
        "optimizer_evaluations": int(result.nfev),
        "completed_at": datetime.now().isoformat(timespec="seconds"),
    }
    atomic_write_json(output, payload)
    return payload


def run_probes(iterations: int, seed: int, force: bool) -> list[dict]:
    PROBE_DIR.mkdir(parents=True, exist_ok=True)
    outputs = []
    pairs = [(uav, missile) for uav in UAV_NAMES for missile in MISSILE_NAMES]
    for index, (uav, missile) in enumerate(pairs, 1):
        log(f"配对探针 {index}/{len(pairs)}: {uav} -> {missile}")
        payload = probe_pair(uav, missile, iterations, seed + index, force)
        outputs.append(payload)
        atomic_write_json(
            STATUS_JSON,
            {
                "stage": "pair_probes",
                "completed": index,
                "total": len(pairs),
                "last_pair": [uav, missile],
                "updated_at": datetime.now().isoformat(timespec="seconds"),
            },
        )
    matrix = {
        uav: {
            missile: next(
                item["exact_duration"]
                for item in outputs
                if item["uav"] == uav and item["missile"] == missile
            )
            for missile in MISSILE_NAMES
        }
        for uav in UAV_NAMES
    }
    atomic_write_json(RUN_DIR / "pair_probe_matrix.json", matrix)
    return outputs


def route_vector(route: list[BombPlan]) -> np.ndarray:
    first = route[0]
    values = [first.heading_deg, first.speed]
    for bomb in sorted(route, key=lambda item: item.bomb_id):
        values.extend([bomb.drop_time, bomb.fuse_delay])
    return np.asarray(values, dtype=float)


def decode_route(template: list[BombPlan], vector: np.ndarray) -> list[BombPlan]:
    ordered = sorted(template, key=lambda item: item.bomb_id)
    heading = float(vector[0] % 360.0)
    speed = float(vector[1])
    result = []
    for i, old in enumerate(ordered):
        result.append(replace(
            old,
            heading_deg=heading,
            speed=speed,
            drop_time=float(vector[2 + 2 * i]),
            fuse_delay=float(vector[3 + 2 * i]),
        ))
    return result


def refine_route(
    route: list[BombPlan],
    base_other: list[BombPlan],
    missile_name: str,
    iterations: int,
    seed: int,
) -> list[BombPlan]:
    """在全局并集目标下精修一架无人机的共享航向/速度与投弹时序。"""
    x0 = route_vector(route)
    n = len(route)
    z0 = UAV_POSITIONS[route[0].uav][2]
    max_tau = min(16.0, math.sqrt(2.0 * z0 / core.G))
    hit_time = MISSILES[missile_name].hit_time
    bounds = [(0.0, 360.0), (70.0, 140.0)]
    for _ in range(n):
        bounds.extend([(0.0, min(50.0, hit_time - 0.31)), (0.3, max_tau)])
    lower = np.asarray([item[0] for item in bounds], dtype=float)
    upper = np.asarray([item[1] for item in bounds], dtype=float)
    # SciPy 对 x0 使用严格边界比较；恰好位于 0.3 的浮点数也统一夹入闭区间。
    epsilon = 1e-10 * np.maximum(1.0, upper - lower)
    x0 = np.minimum(np.maximum(x0, lower + epsilon), upper - epsilon)

    def objective(x) -> float:
        candidate = decode_route(route, np.asarray(x, dtype=float))
        if validate_plan(candidate):
            return 1e4
        all_for_missile = base_other + candidate
        return -fast_score(all_for_missile, missile_name)

    result = differential_evolution(
        objective,
        bounds,
        seed=seed,
        maxiter=max(1, iterations),
        popsize=6,
        tol=1e-7,
        polish=True,
        updating="immediate",
        workers=1,
        x0=x0,
    )
    old_score = -objective(x0)
    new_route = decode_route(route, result.x)
    new_score = -objective(result.x)
    # 若并集没有实质提高，保留原路线，避免把“完全重叠、边际收益为零”的
    # 辅助弹漂移到一个自身也不产生遮蔽的任意参数点。
    if new_score <= old_score + 1e-5 or validate_plan(new_route):
        return route
    log(
        f"{route[0].uav}->{missile_name} 快速并集 "
        f"{old_score:.6f} -> {new_score:.6f} s"
    )
    return new_route


def optimize_plan(iterations: int, seed: int, resume: bool) -> list[BombPlan]:
    if resume and BEST_JSON.exists():
        saved = json.loads(BEST_JSON.read_text(encoding="utf-8"))
        fields = (
            "uav", "missile", "heading_deg", "speed",
            "drop_time", "fuse_delay", "bomb_id",
        )
        start = [
            BombPlan(**{field: item[field] for field in fields})
            for item in saved["bombs"]
        ]
        log("从 q5_best.json 恢复已有策略")
    else:
        start = list(WARM_START)

    by_uav = {uav: [p for p in start if p.uav == uav] for uav in UAV_NAMES}
    # FY1 已由第 3 问高精度精修；重点精修 M2 的三条航线和 FY5->M3。
    for index, uav in enumerate(("FY2", "FY3", "FY4", "FY5"), 1):
        route = by_uav[uav]
        missile_name = route[0].missile
        other = [
            p
            for name, plans in by_uav.items()
            if name != uav
            for p in plans
            if p.missile == missile_name
        ]
        by_uav[uav] = refine_route(
            route,
            other,
            missile_name,
            iterations=iterations,
            seed=seed + index,
        )
        current = [p for name in UAV_NAMES for p in by_uav[name]]
        checkpoint = evaluate_plan(current, exact=True, cooperative=False)
        atomic_write_json(BEST_JSON, checkpoint)
        atomic_write_json(
            STATUS_JSON,
            {
                "stage": "route_refinement",
                "completed": index,
                "total": 4,
                "last_uav": uav,
                "independent_total": checkpoint["independent_total"],
                "updated_at": datetime.now().isoformat(timespec="seconds"),
            },
        )
    return [p for uav in UAV_NAMES for p in by_uav[uav]]


def write_result3_xlsx(result: dict, path: Path) -> None:
    if not TEMPLATE_XLSX.exists():
        raise FileNotFoundError(f"找不到官方模板：{TEMPLATE_XLSX}")
    path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(TEMPLATE_XLSX, path)
    workbook = load_workbook(path)
    sheet = workbook.active
    for row in range(2, 17):
        for column in (2, 3, 5, 6, 7, 8, 9, 10, 11, 12):
            sheet.cell(row, column).value = None

    by_uav: dict[str, list[dict]] = {name: [] for name in UAV_NAMES}
    for bomb in result["bombs"]:
        by_uav[bomb["uav"]].append(bomb)
    row = 2
    for uav in UAV_NAMES:
        bombs = sorted(by_uav[uav], key=lambda item: item["bomb_id"])
        for bomb_id in range(1, 4):
            sheet.cell(row, 1).value = uav
            sheet.cell(row, 4).value = bomb_id
            if bombs:
                sheet.cell(row, 2).value = float(bombs[0]["heading_deg"] % 360.0)
                sheet.cell(row, 3).value = float(bombs[0]["speed"])
            selected = next((b for b in bombs if b["bomb_id"] == bomb_id), None)
            if selected is not None:
                for offset, value in enumerate(selected["drop_point"]):
                    sheet.cell(row, 5 + offset).value = float(value)
                for offset, value in enumerate(selected["explosion_point"]):
                    sheet.cell(row, 8 + offset).value = float(value)
                sheet.cell(row, 11).value = float(selected["assigned_duration"])
                sheet.cell(row, 12).value = selected["missile"]
            row += 1

    for cells in sheet.iter_rows(min_row=2, max_row=16, min_col=2, max_col=11):
        for cell in cells:
            if cell.column == 2:
                cell.number_format = "0.000000"
            elif cell.column == 11:
                cell.number_format = "0.000000"
            else:
                cell.number_format = "0.0000"
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_area = "A1:L18"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(path)


def plot_timeline(result: dict, path: Path) -> None:
    colors = {"M1": "#2563eb", "M2": "#dc2626", "M3": "#16a34a"}
    fig, axes = plt.subplots(3, 1, figsize=(11.5, 7.5), sharex=True)
    for axis, name in zip(axes, MISSILE_NAMES):
        details = result["missiles"][name]
        for a, b in details["independent_intervals"]:
            axis.barh(0, b - a, left=a, height=0.35, color=colors[name], alpha=0.85)
        if "cooperative_intervals" in details:
            for a, b in details["cooperative_intervals"]:
                axis.barh(0.45, b - a, left=a, height=0.22, color="#111827", alpha=0.7)
        axis.set_yticks([0, 0.45] if "cooperative_intervals" in details else [0])
        axis.set_yticklabels(["single-cloud union", "surface coop"] if "cooperative_intervals" in details else ["single-cloud union"])
        axis.set_title(f"{name}: {details['independent_duration']:.6f} s")
        axis.grid(axis="x", alpha=0.25)
    axes[-1].set_xlabel("Time after command (s)")
    fig.suptitle(f"Q5 strict full-cover timeline, total = {result['independent_total']:.6f} s")
    fig.tight_layout()
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def print_summary(result: dict) -> None:
    print("\n==== 第 5 问严格圆柱全遮蔽结果 ====", flush=True)
    for name in MISSILE_NAMES:
        details = result["missiles"][name]
        print(
            f"{name}: {details['independent_duration']:.10f} s  "
            f"intervals={details['independent_intervals']}",
            flush=True,
        )
    print(f"三枚导弹总有效遮蔽时长: {result['independent_total']:.10f} s", flush=True)
    if "cooperative_total" in result:
        print(f"多烟幕完整表面协同复核: {result['cooperative_total']:.10f} s", flush=True)
    print(f"结果表: {RESULT_XLSX}", flush=True)
    print(f"明细 JSON: {BEST_JSON}", flush=True)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--mode",
        choices=("verify", "optimize", "probe"),
        default="verify",
        help="verify 直接终算；optimize 局部精修；probe 跑 15 个配对探针",
    )
    parser.add_argument("--iterations", type=int, default=8, help="每次 DE 的迭代数")
    parser.add_argument("--seed", type=int, default=2025, help="随机种子")
    parser.add_argument("--resume", action="store_true", help="从已有 q5_best.json 恢复")
    parser.add_argument("--force", action="store_true", help="probe 模式重算已有配对")
    parser.add_argument("--cooperative", action="store_true", help="执行多烟幕完整表面协同复核")
    parser.add_argument("--fine-cooperative", action="store_true", help="协同复核使用 1°/0.5m 细网格")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    RUN_DIR.mkdir(parents=True, exist_ok=True)
    started = time.perf_counter()
    log(f"启动 mode={args.mode}, iterations={args.iterations}, seed={args.seed}")
    if args.mode == "probe":
        probes = run_probes(args.iterations, args.seed, args.force)
        positive = [item for item in probes if item["exact_duration"] > 1e-8]
        log(f"15 个探针完成，其中 {len(positive)} 对严格时长为正")
        return
    if args.mode == "verify":
        plans = list(WARM_START)
    else:
        plans = optimize_plan(args.iterations, args.seed, args.resume)
    result = evaluate_plan(
        plans,
        exact=True,
        cooperative=args.cooperative or args.fine_cooperative,
        fine_cooperative=args.fine_cooperative,
    )
    result["run"] = {
        "mode": args.mode,
        "iterations": args.iterations,
        "seed": args.seed,
        "elapsed_seconds": time.perf_counter() - started,
    }
    atomic_write_json(BEST_JSON, result)
    write_result3_xlsx(result, RESULT_XLSX)
    plot_timeline(result, TIMELINE_PNG)
    atomic_write_json(
        STATUS_JSON,
        {
            "stage": "complete",
            "independent_total": result["independent_total"],
            "result_xlsx": str(RESULT_XLSX),
            "best_json": str(BEST_JSON),
            "updated_at": datetime.now().isoformat(timespec="seconds"),
        },
    )
    print_summary(result)
    log(f"完成，总时长 {result['independent_total']:.10f} s")


if __name__ == "__main__":
    main()
