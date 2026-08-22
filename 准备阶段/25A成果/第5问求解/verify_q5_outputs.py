#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""独立复算并核对第 5 问 JSON、result3.xlsx 和时间轴图。"""

from __future__ import annotations

import json
import math
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

from q5_strict_ring import (
    BEST_JSON,
    BombPlan,
    RESULT_XLSX,
    TIMELINE_PNG,
    UAV_NAMES,
    evaluate_plan,
    validate_plan,
)


FIELDS = (
    "uav", "missile", "heading_deg", "speed",
    "drop_time", "fuse_delay", "bomb_id",
)


def close(actual, expected, tolerance=5e-7):
    if not math.isclose(float(actual), float(expected), rel_tol=0.0, abs_tol=tolerance):
        raise AssertionError(f"数值不一致：actual={actual}, expected={expected}")


def main() -> None:
    if not BEST_JSON.exists():
        raise FileNotFoundError(f"缺少 {BEST_JSON}")
    saved = json.loads(BEST_JSON.read_text(encoding="utf-8"))
    plans = [BombPlan(**{field: item[field] for field in FIELDS}) for item in saved["bombs"]]
    errors = validate_plan(plans)
    if errors:
        raise AssertionError("策略约束失败：" + "; ".join(errors))
    recalculated = evaluate_plan(plans, exact=True)
    close(recalculated["independent_total"], saved["independent_total"], 1e-8)
    if recalculated["independent_total"] < 22.0:
        raise AssertionError("严格总时长低于 22 s，疑似结果文件被替换")
    for name in ("M1", "M2", "M3"):
        close(
            recalculated["missiles"][name]["independent_duration"],
            saved["missiles"][name]["independent_duration"],
            1e-8,
        )

    if not RESULT_XLSX.exists():
        raise FileNotFoundError(f"缺少 {RESULT_XLSX}")
    workbook = load_workbook(RESULT_XLSX, data_only=False)
    sheet = workbook.active
    expected_headers = (
        "无人机编号", "无人机运动方向", "无人机运动速度 (m/s)", "烟幕干扰弹编号",
        "烟幕干扰弹投放点的x坐标 (m)", "烟幕干扰弹投放点的y坐标 (m)",
        "烟幕干扰弹投放点的z坐标 (m)", "烟幕干扰弹起爆点的x坐标 (m)",
        "烟幕干扰弹起爆点的y坐标 (m)", "烟幕干扰弹起爆点的z坐标 (m)",
        "有效干扰时长 (s)", "干扰的导弹编号",
    )
    actual_headers = tuple(sheet.cell(1, column).value for column in range(1, 13))
    if actual_headers != expected_headers:
        raise AssertionError("result3.xlsx 表头不再是官方模板表头")

    by_key = {(item["uav"], item["bomb_id"]): item for item in saved["bombs"]}
    for uav_index, uav in enumerate(UAV_NAMES):
        for bomb_id in range(1, 4):
            row = 2 + 3 * uav_index + bomb_id - 1
            if sheet.cell(row, 1).value != uav or sheet.cell(row, 4).value != bomb_id:
                raise AssertionError(f"第 {row} 行无人机/弹号映射错误")
            bomb = by_key.get((uav, bomb_id))
            if bomb is None:
                if any(sheet.cell(row, col).value is not None for col in range(5, 13)):
                    raise AssertionError(f"第 {row} 行未使用烟幕弹不应填写投放数据")
                continue
            close(sheet.cell(row, 2).value, bomb["heading_deg"])
            close(sheet.cell(row, 3).value, bomb["speed"])
            for offset, value in enumerate(bomb["drop_point"]):
                close(sheet.cell(row, 5 + offset).value, value)
            for offset, value in enumerate(bomb["explosion_point"]):
                close(sheet.cell(row, 8 + offset).value, value)
            close(sheet.cell(row, 11).value, bomb["assigned_duration"])
            if sheet.cell(row, 12).value != bomb["missile"]:
                raise AssertionError(f"第 {row} 行导弹编号错误")

    for row in sheet.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.startswith("#"):
                raise AssertionError(f"发现 Excel 错误值 {cell.coordinate}={value}")
    workbook.close()

    if not TIMELINE_PNG.exists():
        raise FileNotFoundError(f"缺少 {TIMELINE_PNG}")
    with Image.open(TIMELINE_PNG) as image:
        if image.width < 1000 or image.height < 700:
            raise AssertionError("时间轴图分辨率异常")

    print(f"[OK] 策略约束全部满足，共 {len(plans)} 枚有效策略弹")
    print(f"[OK] 严格三导弹总时长 {recalculated['independent_total']:.10f} s")
    print("[OK] result3.xlsx 与 JSON 逐行一致，无公式错误")
    print(f"[OK] 时间轴图 {TIMELINE_PNG.name}")


if __name__ == "__main__":
    main()
