#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""核验 Q4 JSON、官方 Excel 输出及物理约束。"""

import json
import math
from pathlib import Path
import sys

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
JSON_PATH = HERE / "q4_runs" / "q4_best.json"
XLSX_PATH = HERE / "result2.xlsx"

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")


def close(a: float, b: float, tol: float = 1e-5) -> bool:
    return math.isclose(float(a), float(b), rel_tol=0.0, abs_tol=tol)


def main() -> None:
    failures: list[str] = []
    if not JSON_PATH.exists():
        failures.append(f"缺少 {JSON_PATH}")
    if not XLSX_PATH.exists() or XLSX_PATH.stat().st_size < 1000:
        failures.append(f"缺少或文件过小：{XLSX_PATH}")
    if failures:
        raise SystemExit("\n".join(failures))

    result = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    uavs = result["uavs"]
    if [uav["name"] for uav in uavs] != ["FY1", "FY2", "FY3"]:
        failures.append("JSON 无人机顺序或数量错误")
    for uav in uavs:
        if not 70.0 <= uav["speed"] <= 140.0:
            failures.append(f"{uav['name']} 速度越界")
        if uav["drop_time"] < -1e-9:
            failures.append(f"{uav['name']} 投放时刻为负")
        if uav["explosion_point"][2] < -1e-9:
            failures.append(f"{uav['name']} 在地下起爆")

    workbook = load_workbook(XLSX_PATH, data_only=False)
    sheet = workbook.active
    for row, uav in enumerate(uavs, 2):
        expected = [
            uav["name"], uav["heading_deg"], uav["speed"],
            *uav["drop_point"], *uav["explosion_point"], uav["duration"],
        ]
        actual = [sheet.cell(row=row, column=column).value for column in range(1, 11)]
        for column, (got, want) in enumerate(zip(actual, expected), 1):
            ok = got == want if column == 1 else got is not None and close(got, want)
            if not ok:
                failures.append(f"Excel {sheet.cell(row=row, column=column).coordinate} 与 JSON 不一致")
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("#"):
                failures.append(f"Excel 错误值：{cell.coordinate}={cell.value}")

    if failures:
        print("[FAIL] Q4 产物核验失败")
        for failure in failures:
            print(" -", failure)
        raise SystemExit(1)
    print("[OK] FY1/FY2/FY3、速度、投放时刻和起爆高度均满足约束")
    print("[OK] result2.xlsx 与 q4_best.json 的 30 个结果单元格一致")
    print("[OK] 工作簿无 Excel 错误值")
    print(f"[OK] 严格并集时长 {result['independent_union_duration']:.10f} s")


if __name__ == "__main__":
    main()
