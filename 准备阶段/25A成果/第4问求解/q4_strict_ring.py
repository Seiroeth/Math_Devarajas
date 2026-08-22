#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""2025 A 题问题 4：FY1/FY2/FY3 各投一弹，对 M1 的严格圆柱遮蔽优化。"""

from __future__ import annotations

import argparse
from datetime import datetime
import importlib
import json
import math
from pathlib import Path
import shutil
import sys
import time

import matplotlib.pyplot as plt
from matplotlib import font_manager
import numpy as np
from openpyxl import load_workbook
from scipy.optimize import differential_evolution, minimize

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
for candidate in [HERE, ROOT, ROOT / "第3问求解", ROOT / "q3_runs" / "第3问求解过程文件"]:
    if (candidate / "smoke_strict_core.py").exists():
        sys.path.insert(0, str(candidate))
        break
else:
    raise FileNotFoundError("未找到 smoke_strict_core.py，请保留第三问公共模型文件。")

core = importlib.import_module("smoke_strict_core")
G = core.G
MISSILE_HIT_TIME = core.MISSILE_HIT_TIME
Smoke = core.Smoke
UAV_POSITIONS = core.UAV_POSITIONS
cooperative_intervals = core.cooperative_intervals
fast_intervals = core.fast_intervals
independent_union = core.independent_union
interval_length = core.interval_length

PROJECT_ROOT = HERE.parents[2]
TEMPLATE = PROJECT_ROOT / "数学建模学习资料" / "历年赛题题目" / "2025A" / "附件" / "result2.xlsx"
RUN_DIR = HERE / "q4_runs"
LOG_DIR = RUN_DIR / "logs"
CHECKPOINT_DIR = RUN_DIR / "checkpoints"
RESULT_JSON = RUN_DIR / "q4_best.json"
RESULT_XLSX = HERE / "result2.xlsx"

UAV_NAMES = ("FY1", "FY2", "FY3")

# 网页思路引用的开源候选，仅用于 benchmark。
REFERENCE = np.array([
    176.6188, 70.0, 0.0, 2.4840,
    306.1909, 136.85, 8.551, 3.9860,
    122.4743, 92.96, 31.733, 7.6460,
])

# 在本项目连续圆周判据下逐坐标严格精修后的候选。
REFINED = np.array([
    176.6430, 70.0, 0.0, 2.49705,
    298.5250, 139.77, 6.457245512, 4.696824240,
    106.0800, 93.749854590, 28.212156240, 6.0950,
])

QUICK_BOUNDS = {
    "FY1": [(165.0, 190.0), (70.0, 100.0), (0.0, 5.0), (1.0, 5.0)],
    "FY2": [(280.0, 330.0), (100.0, 140.0), (3.0, 20.0), (1.0, 8.0)],
    "FY3": [(95.0, 150.0), (70.0, 125.0), (18.0, 48.0), (2.0, 11.8)],
}

SEARCH = {
    "quick": {"dt": 0.08, "nphi": 72, "maxiter": 35, "popsize": 8, "tol": 2e-5},
    "full": {"dt": 0.04, "nphi": 180, "maxiter": 120, "popsize": 14, "tol": 2e-7},
}

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


def write_json(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_suffix(path.suffix + ".tmp")
    temp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    temp.replace(path)


def decode(x: np.ndarray) -> list[Smoke]:
    x = np.asarray(x, dtype=float).reshape(3, 4)
    return [
        Smoke(math.radians(row[0] % 360.0), row[1], row[2], row[3], UAV_POSITIONS[name])
        for name, row in zip(UAV_NAMES, x)
    ]


def single_smoke(name: str, x: np.ndarray) -> Smoke:
    psi_deg, speed, td, tau = map(float, x)
    return Smoke(math.radians(psi_deg % 360.0), speed, td, tau, UAV_POSITIONS[name])


def fast_single_duration(name: str, x: np.ndarray, dt: float, nphi: int) -> float:
    smoke = single_smoke(name, x)
    return interval_length(fast_intervals(smoke, dt=dt, nphi=nphi)) if smoke.valid else 0.0


def fast_union_duration(x: np.ndarray, dt: float, nphi: int) -> float:
    smokes = decode(x)
    if not all(smoke.valid for smoke in smokes):
        return 0.0
    return independent_union(smokes, exact=False, dt=dt, nphi=nphi)[2]


def full_bounds(name: str) -> list[tuple[float, float]]:
    tau_max = math.sqrt(2.0 * UAV_POSITIONS[name][2] / G)
    return [(0.0, 360.0), (70.0, 140.0), (0.0, 55.0), (0.05, tau_max)]


class SingleTracker:
    def __init__(self, name: str, mode: str, dt: float, nphi: int, start: np.ndarray):
        self.name = name
        self.mode = mode
        self.dt = dt
        self.nphi = nphi
        self.generation = 0
        self.started = time.time()
        self.best_x = np.asarray(start, dtype=float).copy()
        self.best_duration = fast_single_duration(name, start, dt, nphi)

    def objective(self, x: np.ndarray) -> float:
        duration = fast_single_duration(self.name, x, self.dt, self.nphi)
        if duration > self.best_duration:
            self.best_duration = duration
            self.best_x = np.asarray(x, dtype=float).copy()
        return -duration

    def callback(self, xk: np.ndarray, convergence: float) -> bool:
        self.generation += 1
        duration = fast_single_duration(self.name, xk, self.dt, self.nphi)
        if duration > self.best_duration:
            self.best_duration = duration
            self.best_x = np.asarray(xk, dtype=float).copy()
        status = {
            "uav": self.name,
            "mode": self.mode,
            "generation": self.generation,
            "best_duration_fast": self.best_duration,
            "best_x": self.best_x.tolist(),
            "convergence": float(convergence),
            "elapsed_seconds": time.time() - self.started,
            "updated_at": datetime.now().isoformat(timespec="seconds"),
        }
        write_json(LOG_DIR / "status.json", status)
        write_json(CHECKPOINT_DIR / f"{self.mode}_{self.name}_latest.json", status)
        print(
            f"{self.name} generation={self.generation:03d}  "
            f"fast={self.best_duration:.6f} s  conv={convergence:.3e}",
            flush=True,
        )
        return False


def optimize_single(name: str, start: np.ndarray, mode: str, force: bool) -> np.ndarray:
    checkpoint = CHECKPOINT_DIR / f"{mode}_{name}_complete.json"
    refined = REFINED.reshape(3, 4)[UAV_NAMES.index(name)]
    if checkpoint.exists() and not force:
        saved = np.asarray(json.loads(checkpoint.read_text(encoding="utf-8"))["x"], dtype=float)
        print(f"读取 {name} 的 {mode} 完成检查点。")
        return max(
            [saved, refined],
            key=lambda x: fast_single_duration(name, x, dt=0.02, nphi=360),
        )

    settings = SEARCH[mode]
    bounds = QUICK_BOUNDS[name] if mode == "quick" else full_bounds(name)
    tracker = SingleTracker(name, mode, settings["dt"], settings["nphi"], start)
    result = differential_evolution(
        tracker.objective,
        bounds=bounds,
        seed=2025 + UAV_NAMES.index(name),
        popsize=settings["popsize"],
        maxiter=settings["maxiter"],
        tol=settings["tol"],
        polish=False,
        updating="immediate",
        workers=1,
        x0=start,
        callback=tracker.callback,
    )
    candidates = [start, refined, tracker.best_x, np.asarray(result.x)]
    local = minimize(
        tracker.objective,
        max(candidates, key=lambda x: fast_single_duration(name, x, settings["dt"], settings["nphi"])),
        method="Powell",
        bounds=bounds,
        options={"maxiter": 160 if mode == "full" else 70, "xtol": 1e-5, "ftol": 1e-6},
    )
    candidates.append(np.asarray(local.x))
    best = max(candidates, key=lambda x: fast_single_duration(name, x, dt=0.02, nphi=360))
    write_json(checkpoint, {
        "uav": name,
        "mode": mode,
        "x": best.tolist(),
        "fast_duration_fine": fast_single_duration(name, best, dt=0.02, nphi=360),
        "completed_at": datetime.now().isoformat(timespec="seconds"),
    })
    return best


def optimize(mode: str, force: bool = False) -> np.ndarray:
    starts = REFINED.reshape(3, 4)
    rows = [optimize_single(name, starts[i], mode, force) for i, name in enumerate(UAV_NAMES)]
    candidate = np.concatenate(rows)
    return max(
        [candidate, REFINED],
        key=lambda x: fast_union_duration(x, dt=0.02, nphi=360),
    )


def exact_report(x: np.ndarray, cooperative: bool = True) -> dict:
    smokes = decode(x)
    each, union, total = independent_union(smokes, exact=True)
    result = {
        "x": np.asarray(x, dtype=float).tolist(),
        "missile_hit_time": MISSILE_HIT_TIME,
        "uavs": [],
        "independent_union_intervals": [[a, b] for a, b in union],
        "independent_union_duration": total,
    }
    for name, smoke, intervals in zip(UAV_NAMES, smokes, each):
        result["uavs"].append({
            "name": name,
            "heading_deg": math.degrees(smoke.psi) % 360.0,
            "speed": smoke.speed,
            "drop_time": smoke.td,
            "fuse_delay": smoke.tau,
            "explosion_time": smoke.te,
            "drop_point": smoke.drop_point.tolist(),
            "explosion_point": smoke.explosion_point.tolist(),
            "intervals": [[a, b] for a, b in intervals],
            "duration": interval_length(intervals),
        })
    if cooperative:
        coop, point_count = cooperative_intervals(
            smokes,
            time_dt=0.02,
            theta_step_deg=1.0,
            z_step=0.5,
            radial_step=0.5,
        )
        coop_total = interval_length(coop)
        result.update({
            "cooperative_surface_points": point_count,
            "cooperative_intervals": [[a, b] for a, b in coop],
            "cooperative_duration": coop_total,
            "cooperative_gain": coop_total - total,
        })
    result["verified_at"] = datetime.now().isoformat(timespec="seconds")
    return result


def export_xlsx(result: dict, path: Path = RESULT_XLSX) -> None:
    if not TEMPLATE.exists():
        raise FileNotFoundError(f"未找到官方模板：{TEMPLATE}")
    shutil.copy2(TEMPLATE, path)
    workbook = load_workbook(path)
    sheet = workbook.active
    for row, uav in enumerate(result["uavs"], 2):
        values = [
            uav["name"], uav["heading_deg"], uav["speed"],
            *uav["drop_point"], *uav["explosion_point"], uav["duration"],
        ]
        for column, value in enumerate(values, 1):
            sheet.cell(row=row, column=column, value=value if column == 1 else float(value))
        for column in range(2, 11):
            sheet.cell(row=row, column=column).number_format = "0.000000"
    sheet.print_area = "A1:J6"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.35
    sheet.page_margins.bottom = 0.35
    workbook.save(path)


def plot_timeline(result: dict, path: Path) -> None:
    font_path = font_manager.findfont("Microsoft YaHei")
    plt.rcParams["font.family"] = font_manager.FontProperties(fname=font_path).get_name()
    plt.rcParams["axes.unicode_minus"] = False
    fig, ax = plt.subplots(figsize=(11, 4.0))
    colors = ["#4472C4", "#ED7D31", "#70AD47"]
    for i, uav in enumerate(result["uavs"]):
        for a, b in uav["intervals"]:
            ax.barh(i + 1, b - a, left=a, height=0.55, color=colors[i])
    for a, b in result["independent_union_intervals"]:
        ax.barh(0, b - a, left=a, height=0.55, color="#7030A0")
    ax.set_yticks([0, 1, 2, 3], ["并集", "FY1", "FY2", "FY3"])
    ax.set_xlabel("任务开始后的时间 / s")
    ax.set_title(f"Q4 严格圆柱遮蔽时间轴（并集 {result['independent_union_duration']:.6f} s）")
    ax.grid(axis="x", alpha=0.25)
    fig.tight_layout()
    fig.savefig(path, dpi=180)
    plt.close(fig)


def print_report(result: dict) -> None:
    print("=== Q4 严格圆柱判据 ===")
    for uav in result["uavs"]:
        print(
            f"{uav['name']}: heading={uav['heading_deg']:.9f} deg, "
            f"speed={uav['speed']:.9f}, td={uav['drop_time']:.9f}, "
            f"tau={uav['fuse_delay']:.9f}, te={uav['explosion_time']:.9f}, "
            f"duration={uav['duration']:.9f}"
        )
        print("  投放点:", np.array2string(np.asarray(uav["drop_point"]), precision=6))
        print("  起爆点:", np.array2string(np.asarray(uav["explosion_point"]), precision=6))
        print("  区间  :", uav["intervals"])
    print("独立完整遮蔽并集:", result["independent_union_intervals"])
    print(f"独立完整遮蔽总时长: {result['independent_union_duration']:.10f} s")
    if "cooperative_duration" in result:
        print(f"表面 max-min 联合时长: {result['cooperative_duration']:.10f} s")
        print(f"协同增益: {result['cooperative_gain']:.10f} s")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--mode", choices=["verify", "quick", "full"], default="verify")
    parser.add_argument("--force", action="store_true", help="忽略完成检查点并重跑搜索")
    parser.add_argument("--skip-coop", action="store_true", help="跳过完整表面多烟幕精检")
    args = parser.parse_args()

    LOG_DIR.mkdir(parents=True, exist_ok=True)
    CHECKPOINT_DIR.mkdir(parents=True, exist_ok=True)
    x = REFINED.copy() if args.mode == "verify" else optimize(args.mode, args.force)
    result = exact_report(x, cooperative=not args.skip_coop)
    write_json(RESULT_JSON, result)
    export_xlsx(result)
    plot_timeline(result, HERE / "q4_timeline.png")
    print_report(result)
    print(f"JSON : {RESULT_JSON}")
    print(f"Excel: {RESULT_XLSX}")


if __name__ == "__main__":
    main()
